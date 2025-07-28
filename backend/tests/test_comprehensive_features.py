#!/usr/bin/env python3
"""
Comprehensive Test Script for File Upload and Processing System

This script demonstrates all the implemented features:
- File upload with virus scanning
- Background processing with Celery
- Partial processing for problematic files
- Data recovery mechanisms
- Manual intervention tools
- Analytics and monitoring
- Cloud storage integration
- File cleanup and retention
"""

import pytest
import asyncio
import json
import tempfile
import os
from pathlib import Path
from datetime import datetime
import requests
import time

# Mark all tests in this file as integration tests
pytestmark = pytest.mark.integration

# Test configuration
API_BASE = "http://localhost:8000/api/v1"
TEST_USER = {
    "username": "test_user",
    "email": "test@example.com",
    "password": "testpassword123",
}
ADMIN_USER = {
    "username": "admin_user",
    "email": "admin@example.com",
    "password": "adminpassword123",
}


class FinVisionTester:
    """Comprehensive tester for FinVision system."""

    def __init__(self):
        self.user_token = None
        self.admin_token = None
        self.test_file_id = None

    def authenticate(self):
        """Authenticate test users."""
        print("🔐 Authenticating users...")

        # Authenticate regular user
        try:
            response = requests.post(
                f"{API_BASE}/auth/login",
                data={
                    "username": TEST_USER["username"],
                    "password": TEST_USER["password"],
                },
            )
            if response.status_code == 200:
                self.user_token = response.json()["access_token"]
                print(f"✅ User authenticated: {TEST_USER['username']}")
            else:
                print(f"❌ User authentication failed: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ User authentication error: {e}")
            return False

        # Authenticate admin user
        try:
            response = requests.post(
                f"{API_BASE}/auth/login",
                data={
                    "username": ADMIN_USER["username"],
                    "password": ADMIN_USER["password"],
                },
            )
            if response.status_code == 200:
                self.admin_token = response.json()["access_token"]
                print(f"✅ Admin authenticated: {ADMIN_USER['username']}")
                return True
            else:
                print(f"❌ Admin authentication failed: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Admin authentication error: {e}")
            return False

    def create_test_excel_file(self) -> str:
        """Create a test Excel file for upload."""
        print("📊 Creating test Excel file...")

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            # Create workbook
            wb = openpyxl.Workbook()

            # P&L Sheet
            pnl_sheet = wb.active
            pnl_sheet.title = "P&L Statement"

            # Add P&L data
            pnl_data = [
                ["P&L Statement", "", "2023"],
                ["", "", ""],
                ["Revenue", "", 1000000],
                ["Cost of Sales", "", 600000],
                ["Gross Profit", "", "=C3-C4"],
                ["", "", ""],
                ["Operating Expenses", "", ""],
                ["Salaries", "", 200000],
                ["Rent", "", 50000],
                ["Marketing", "", 75000],
                ["Total OpEx", "", "=C8+C9+C10"],
                ["", "", ""],
                ["Operating Profit", "", "=C5-C11"],
                ["", "", ""],
                ["Net Income", "", "=C13"],
            ]

            for row_num, row_data in enumerate(pnl_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = pnl_sheet.cell(row=row_num, column=col_num)
                    cell.value = value

            # Balance Sheet
            bs_sheet = wb.create_sheet("Balance Sheet")
            bs_data = [
                ["Balance Sheet", "", "2023"],
                ["", "", ""],
                ["ASSETS", "", ""],
                ["Current Assets", "", ""],
                ["Cash", "", 250000],
                ["Accounts Receivable", "", 150000],
                ["Inventory", "", 100000],
                ["Total Current Assets", "", "=C5+C6+C7"],
                ["", "", ""],
                ["Fixed Assets", "", 500000],
                ["Total Assets", "", "=C8+C10"],
                ["", "", ""],
                ["LIABILITIES", "", ""],
                ["Current Liabilities", "", 200000],
                ["Long-term Debt", "", 300000],
                ["Total Liabilities", "", "=C14+C15"],
                ["", "", ""],
                ["EQUITY", "", ""],
                ["Shareholders Equity", "", "=C11-C16"],
                ["Total Liab + Equity", "", "=C16+C19"],
            ]

            for row_num, row_data in enumerate(bs_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = bs_sheet.cell(row=row_num, column=col_num)
                    cell.value = value

            # Cash Flow Sheet
            cf_sheet = wb.create_sheet("Cash Flow")
            cf_data = [
                ["Cash Flow Statement", "", "2023"],
                ["", "", ""],
                ["Operating Activities", "", ""],
                ["Net Income", "", 150000],
                ["Depreciation", "", 25000],
                ["Changes in Working Capital", "", -50000],
                ["Net Cash from Operations", "", "=C4+C5+C6"],
                ["", "", ""],
                ["Investing Activities", "", ""],
                ["Capital Expenditures", "", -100000],
                ["Net Cash from Investing", "", "=C10"],
                ["", "", ""],
                ["Financing Activities", "", ""],
                ["Debt Issuance", "", 50000],
                ["Net Cash from Financing", "", "=C14"],
                ["", "", ""],
                ["Net Change in Cash", "", "=C7+C11+C15"],
            ]

            for row_num, row_data in enumerate(cf_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = cf_sheet.cell(row=row_num, column=col_num)
                    cell.value = value

            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(temp_file.name)
            temp_file.close()

            print(f"✅ Test Excel file created: {temp_file.name}")
            return temp_file.name

        except Exception as e:
            print(f"❌ Failed to create test Excel file: {e}")
            return None

    def test_file_upload(self, file_path: str):
        """Test file upload functionality."""
        print("📤 Testing file upload...")

        try:
            headers = {"Authorization": f"Bearer {self.user_token}"}

            with open(file_path, "rb") as f:
                files = {
                    "file": (
                        "test_financial_model.xlsx",
                        f,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                }

                response = requests.post(
                    f"{API_BASE}/files/upload", files=files, headers=headers
                )

            if response.status_code == 200:
                result = response.json()
                self.test_file_id = result["file_id"]
                print(f"✅ File uploaded successfully: ID {self.test_file_id}")
                print(f"   File size: {result['file_size']} bytes")
                print(
                    f"   Virus scan: {'Clean' if result['virus_scan_clean'] else 'Threats detected'}"
                )
                return True
            else:
                print(f"❌ File upload failed: {response.status_code}")
                print(f"   Response: {response.text}")
                return False

        except Exception as e:
            print(f"❌ File upload error: {e}")
            return False

    def test_file_processing(self):
        """Test background file processing."""
        print("⚙️ Testing file processing...")

        if not self.test_file_id:
            print("❌ No file ID available for processing test")
            return False

        try:
            headers = {"Authorization": f"Bearer {self.user_token}"}

            # Trigger processing
            response = requests.post(
                f"{API_BASE}/files/{self.test_file_id}/process",
                json={"processing_options": {"extract_formulas": True}},
                headers=headers,
            )

            if response.status_code == 200:
                result = response.json()
                task_id = result.get("task_id")
                print(f"✅ Processing started: Task ID {task_id}")

                # Monitor processing status
                if task_id:
                    return self.monitor_processing_status(task_id)
                else:
                    return True
            else:
                print(f"❌ Processing failed to start: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ Processing error: {e}")
            return False

    def monitor_processing_status(self, task_id: str, max_wait: int = 60):
        """Monitor processing status."""
        print(f"👁️ Monitoring processing status for task {task_id}...")

        headers = {"Authorization": f"Bearer {self.user_token}"}
        start_time = time.time()

        while time.time() - start_time < max_wait:
            try:
                response = requests.get(
                    f"{API_BASE}/files/task/{task_id}/status", headers=headers
                )

                if response.status_code == 200:
                    status = response.json()
                    state = status.get("state", "UNKNOWN")
                    progress = status.get("current", 0)
                    total = status.get("total", 100)

                    print(f"   Status: {state} ({progress}/{total})")

                    if state == "SUCCESS":
                        print("✅ Processing completed successfully!")
                        return True
                    elif state == "FAILURE":
                        print("❌ Processing failed!")
                        print(f"   Error: {status.get('error', 'Unknown error')}")
                        return False

                    time.sleep(2)
                else:
                    print(f"❌ Status check failed: {response.status_code}")
                    return False

            except Exception as e:
                print(f"❌ Status monitoring error: {e}")
                return False

        print("⏰ Processing status monitoring timed out")
        return False

    def test_analytics_dashboard(self):
        """Test analytics dashboard functionality."""
        print("📊 Testing analytics dashboard...")

        try:
            headers = {"Authorization": f"Bearer {self.user_token}"}

            # Get dashboard summary
            response = requests.get(
                f"{API_BASE}/analytics/dashboard?days=7", headers=headers
            )

            if response.status_code == 200:
                data = response.json()
                overview = data.get("overview", {})

                print("✅ Analytics dashboard loaded successfully!")
                print(f"   Total files: {overview.get('total_files', 0)}")
                print(f"   Success rate: {overview.get('success_rate', 0)}%")
                print(
                    f"   Avg processing time: {overview.get('average_processing_time_minutes', 0)} min"
                )

                return True
            else:
                print(f"❌ Analytics dashboard failed: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ Analytics error: {e}")
            return False

    def test_partial_processing(self):
        """Test partial processing functionality."""
        print("🔧 Testing partial processing...")

        if not self.test_file_id:
            print("❌ No file ID available for partial processing test")
            return False

        try:
            headers = {"Authorization": f"Bearer {self.user_token}"}

            # Analyze partial processing potential
            response = requests.post(
                f"{API_BASE}/admin-tools/partial-processing/analyze",
                json={
                    "file_id": self.test_file_id,
                    "ignore_validation_errors": True,
                    "extract_available_data": True,
                },
                headers=headers,
            )

            if response.status_code == 200:
                result = response.json()
                print(f"✅ Partial processing analysis completed!")
                print(
                    f"   Can partial process: {result.get('can_partial_process', False)}"
                )
                print(
                    f"   Estimated completion: {result.get('estimated_completion', 0)}%"
                )
                print(f"   Extractable sheets: {result.get('extractable_sheets', [])}")

                return True
            else:
                print(f"❌ Partial processing analysis failed: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ Partial processing error: {e}")
            return False

    def test_data_recovery(self):
        """Test data recovery functionality."""
        print("🚑 Testing data recovery...")

        if not self.test_file_id:
            print("❌ No file ID available for recovery test")
            return False

        try:
            headers = {"Authorization": f"Bearer {self.user_token}"}

            # Get recovery options
            response = requests.get(
                f"{API_BASE}/admin-tools/recovery/{self.test_file_id}/options",
                headers=headers,
            )

            if response.status_code == 200:
                result = response.json()
                options = result.get("recovery_options", [])

                print(f"✅ Recovery options analyzed!")
                print(f"   Available options: {len(options)}")

                for i, option in enumerate(options[:3]):  # Show first 3 options
                    print(f"   Option {i+1}: {option['description']}")
                    print(
                        f"     Success rate: {option['estimated_success_rate']*100:.1f}%"
                    )

                return True
            else:
                print(f"❌ Recovery analysis failed: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ Recovery error: {e}")
            return False

    def test_manual_intervention(self):
        """Test manual intervention functionality."""
        print("🛠️ Testing manual intervention...")

        if not self.test_file_id or not self.admin_token:
            print("❌ Missing file ID or admin token for intervention test")
            return False

        try:
            headers = {"Authorization": f"Bearer {self.admin_token}"}

            # Request manual intervention
            response = requests.post(
                f"{API_BASE}/admin-tools/interventions/request",
                json={
                    "intervention_type": "status_override",
                    "file_id": self.test_file_id,
                    "priority": "normal",
                    "description": "Test status override for demonstration",
                    "justification": "Automated testing of intervention system",
                    "parameters": {"new_status": "completed"},
                },
                headers=headers,
            )

            if response.status_code == 200:
                result = response.json()
                print(f"✅ Manual intervention requested!")
                print(f"   Request ID: {result.get('request_id')}")

                # Check pending interventions
                response = requests.get(
                    f"{API_BASE}/admin-tools/interventions/pending", headers=headers
                )

                if response.status_code == 200:
                    pending = response.json()
                    print(
                        f"   Pending interventions: {pending.get('total_pending', 0)}"
                    )

                return True
            else:
                print(f"❌ Manual intervention failed: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ Manual intervention error: {e}")
            return False

    def test_system_health(self):
        """Test system health monitoring."""
        print("🏥 Testing system health monitoring...")

        if not self.admin_token:
            print("❌ Missing admin token for health test")
            return False

        try:
            headers = {"Authorization": f"Bearer {self.admin_token}"}

            # Get system health
            response = requests.get(
                f"{API_BASE}/admin-tools/system/health", headers=headers
            )

            if response.status_code == 200:
                health = response.json()
                print(f"✅ System health check completed!")
                print(f"   Overall status: {health.get('overall_status', 'unknown')}")

                services = health.get("services", {})
                for service, status in services.items():
                    service_status = (
                        status.get("status", "unknown")
                        if isinstance(status, dict)
                        else status
                    )
                    print(f"   {service}: {service_status}")

                return True
            else:
                print(f"❌ System health check failed: {response.status_code}")
                return False

        except Exception as e:
            print(f"❌ System health error: {e}")
            return False

    def cleanup_test_file(self, file_path: str):
        """Clean up test files."""
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
                print(f"🧹 Cleaned up test file: {file_path}")
        except Exception as e:
            print(f"❌ Cleanup error: {e}")

    def run_comprehensive_test(self):
        """Run all tests in sequence."""
        print("🚀 Starting comprehensive FinVision test suite...")
        print("=" * 60)

        # Create test file
        test_file = self.create_test_excel_file()
        if not test_file:
            print("❌ Failed to create test file. Aborting tests.")
            return False

        try:
            # Test sequence
            tests = [
                ("Authentication", self.authenticate),
                ("File Upload", lambda: self.test_file_upload(test_file)),
                ("File Processing", self.test_file_processing),
                ("Analytics Dashboard", self.test_analytics_dashboard),
                ("Partial Processing", self.test_partial_processing),
                ("Data Recovery", self.test_data_recovery),
                ("Manual Intervention", self.test_manual_intervention),
                ("System Health", self.test_system_health),
            ]

            results = {}

            for test_name, test_func in tests:
                print(f"\n📋 Running: {test_name}")
                print("-" * 40)

                try:
                    result = test_func()
                    results[test_name] = result

                    if result:
                        print(f"✅ {test_name}: PASSED")
                    else:
                        print(f"❌ {test_name}: FAILED")

                except Exception as e:
                    print(f"❌ {test_name}: ERROR - {e}")
                    results[test_name] = False

                time.sleep(1)  # Brief pause between tests

            # Summary
            print("\n" + "=" * 60)
            print("📊 TEST SUMMARY")
            print("=" * 60)

            passed = sum(1 for result in results.values() if result)
            total = len(results)

            for test_name, result in results.items():
                status = "✅ PASSED" if result else "❌ FAILED"
                print(f"{test_name}: {status}")

            print(f"\nTotal: {passed}/{total} tests passed ({passed/total*100:.1f}%)")

            if passed == total:
                print("\n🎉 ALL TESTS PASSED! FinVision system is working correctly.")
            else:
                print(
                    f"\n⚠️ {total-passed} tests failed. Please review the issues above."
                )

            return passed == total

        finally:
            # Cleanup
            self.cleanup_test_file(test_file)


def main():
    """Main test runner."""
    print("FinVision Comprehensive Test Suite")
    print(
        "This script tests all major functionality of the file upload and processing system."
    )
    print("\nEnsure the following services are running:")
    print("- FastAPI server (uvicorn main:app --reload)")
    print("- Redis server")
    print("- Celery worker (celery -A app.core.celery_app worker --loglevel=info)")
    print("- Database is set up and running")

    input("\nPress Enter to continue with tests, or Ctrl+C to abort...")

    tester = FinVisionTester()
    success = tester.run_comprehensive_test()

    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
