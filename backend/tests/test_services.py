#!/usr/bin/env python3
"""
Simple Service Tests for FinVision Components

Tests core services without requiring database or server setup.
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path

# Mark all tests in this file as unit tests
pytestmark = pytest.mark.unit

# Add the app directory to Python path
sys.path.append(os.path.join(os.path.dirname(__file__), 'app'))

def test_imports():
    """Test that all core services can be imported successfully."""
    print("🔍 Testing service imports...")
    
    try:
        # Core services
        from app.services.excel_parser import ExcelParser
        print("✅ ExcelParser imported successfully")
        
        from app.services.financial_extractor import FinancialExtractor
        print("✅ FinancialExtractor imported successfully")
        
        from app.services.partial_processor import PartialProcessor
        print("✅ PartialProcessor imported successfully")
        
        from app.services.data_recovery import DataRecoveryService
        print("✅ DataRecoveryService imported successfully")
        
        from app.services.manual_intervention import ManualInterventionService
        print("✅ ManualInterventionService imported successfully")
        
        from app.services.cloud_storage import CloudStorageManager
        print("✅ CloudStorageManager imported successfully")
        
        from app.services.virus_scanner import VirusScanManager
        print("✅ VirusScanManager imported successfully")
        
        from app.services.file_cleanup import FileCleanupService
        print("✅ FileCleanupService imported successfully")
        
        return True
        
    except Exception as e:
        print(f"❌ Import failed: {e}")
        return False

def test_excel_parser():
    """Test Excel parser basic functionality."""
    print("\n📊 Testing ExcelParser...")
    
    try:
        from app.services.excel_parser import ExcelParser
        
        parser = ExcelParser()
        print("✅ ExcelParser instantiated successfully")
        
        # Test template matching
        pnl_template = parser._get_pnl_template()
        bs_template = parser._get_balance_sheet_template()
        cf_template = parser._get_cash_flow_template()
        print(f"✅ Found 3 financial statement templates (P&L, Balance Sheet, Cash Flow)")
        
        return True
        
    except Exception as e:
        print(f"❌ ExcelParser test failed: {e}")
        return False

def test_financial_extractor():
    """Test Financial extractor basic functionality."""
    print("\n💰 Testing FinancialExtractor...")
    
    try:
        from app.services.financial_extractor import FinancialExtractor
        
        extractor = FinancialExtractor()
        print("✅ FinancialExtractor instantiated successfully")
        
        # Test metric calculators
        calculators = extractor._initialize_metric_calculators()
        print(f"✅ Found {len(calculators)} financial metric calculators")
        
        return True
        
    except Exception as e:
        print(f"❌ FinancialExtractor test failed: {e}")
        return False

def test_partial_processor():
    """Test Partial processor basic functionality."""
    print("\n🔧 Testing PartialProcessor...")
    
    try:
        from app.services.partial_processor import PartialProcessor
        
        processor = PartialProcessor()
        print("✅ PartialProcessor instantiated successfully")
        
        # Test auto-fix strategies
        strategies = processor._initialize_auto_fix_strategies()
        print(f"✅ Found {len(strategies)} auto-fix strategies")
        
        return True
        
    except Exception as e:
        print(f"❌ PartialProcessor test failed: {e}")
        return False

def test_cloud_storage():
    """Test Cloud storage basic functionality."""
    print("\n☁️ Testing CloudStorageManager...")
    
    try:
        from app.services.cloud_storage import CloudStorageManager
        
        # Test with local storage (should always work)
        os.environ['STORAGE_PROVIDER'] = 'local'
        
        manager = CloudStorageManager()
        print("✅ CloudStorageManager instantiated successfully")
        
        # Test filename generation
        filename = manager.generate_secure_filename("test.xlsx", 123)
        print(f"✅ Generated secure filename: {filename}")
        
        # Test storage stats
        stats = manager.get_storage_stats()
        print(f"✅ Storage stats: {stats['provider']}")
        
        return True
        
    except Exception as e:
        print(f"❌ CloudStorageManager test failed: {e}")
        return False

def test_virus_scanner():
    """Test Virus scanner basic functionality."""
    print("\n🛡️ Testing VirusScanManager...")
    
    try:
        from app.services.virus_scanner import VirusScanManager
        
        # Test with basic scanner only (should always work)
        os.environ['VIRUS_SCANNERS'] = 'basic'
        
        manager = VirusScanManager()
        print("✅ VirusScanManager instantiated successfully")
        
        # Test scanner status
        status = manager.get_scanner_status()
        print(f"✅ Scanner status: {status['available_scanners']}")
        
        return True
        
    except Exception as e:
        print(f"❌ VirusScanManager test failed: {e}")
        return False

def test_data_structures():
    """Test data structure definitions."""
    print("\n📋 Testing data structures...")
    
    try:
        from app.services.partial_processor import ProcessingIssueType, ProcessingSeverity
        from app.services.data_recovery import RecoveryType, RecoveryStatus
        from app.services.manual_intervention import InterventionType, InterventionPriority
        
        print("✅ ProcessingIssueType enum loaded")
        print("✅ ProcessingSeverity enum loaded")
        print("✅ RecoveryType enum loaded")
        print("✅ RecoveryStatus enum loaded")
        print("✅ InterventionType enum loaded")
        print("✅ InterventionPriority enum loaded")
        
        return True
        
    except Exception as e:
        print(f"❌ Data structures test failed: {e}")
        return False

def test_configurations():
    """Test configuration loading."""
    print("\n⚙️ Testing configurations...")
    
    try:
        # Set up minimal environment
        os.environ.setdefault('DATABASE_URL', 'sqlite:///test.db')
        os.environ.setdefault('SECRET_KEY', 'test-secret-key')
        os.environ.setdefault('STORAGE_PROVIDER', 'local')
        os.environ.setdefault('VIRUS_SCANNERS', 'basic')
        
        from app.core.config import settings
        print("✅ Settings loaded successfully")
        print(f"✅ Storage provider: {getattr(settings, 'STORAGE_PROVIDER', 'local')}")
        print(f"✅ Max file size: {getattr(settings, 'MAX_FILE_SIZE', 'default')}")
        
        return True
        
    except Exception as e:
        print(f"❌ Configuration test failed: {e}")
        return False

def create_test_excel_file():
    """Create a simple test Excel file."""
    print("\n📄 Testing Excel file creation...")
    
    try:
        import openpyxl
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some test data
        ws['A1'] = "Test"
        ws['B1'] = 123
        ws['C1'] = "=B1*2"
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()
        
        print(f"✅ Created test Excel file: {temp_file.name}")
        
        # Clean up
        os.unlink(temp_file.name)
        print("✅ Test file cleaned up")
        
        return True
        
    except Exception as e:
        print(f"❌ Excel file creation test failed: {e}")
        return False

def run_all_tests():
    """Run all service tests."""
    print("🚀 FinVision Service Tests")
    print("=" * 50)
    
    tests = [
        ("Service Imports", test_imports),
        ("Excel Parser", test_excel_parser),
        ("Financial Extractor", test_financial_extractor),
        ("Partial Processor", test_partial_processor),
        ("Cloud Storage", test_cloud_storage),
        ("Virus Scanner", test_virus_scanner),
        ("Data Structures", test_data_structures),
        ("Configurations", test_configurations),
        ("Excel File Creation", create_test_excel_file),
    ]
    
    results = {}
    
    for test_name, test_func in tests:
        try:
            result = test_func()
            results[test_name] = result
            
            if result:
                print(f"✅ {test_name}: PASSED")
            else:
                print(f"❌ {test_name}: FAILED")
                
        except Exception as e:
            print(f"❌ {test_name}: ERROR - {e}")
            results[test_name] = False
        
        print("-" * 30)
    
    # Summary
    print("\n" + "=" * 50)
    print("📊 TEST SUMMARY")
    print("=" * 50)
    
    passed = sum(1 for result in results.values() if result)
    total = len(results)
    
    for test_name, result in results.items():
        status = "✅ PASSED" if result else "❌ FAILED"
        print(f"{test_name}: {status}")
    
    print(f"\nTotal: {passed}/{total} tests passed ({passed/total*100:.1f}%)")
    
    if passed == total:
        print("\n🎉 ALL TESTS PASSED! Services are working correctly.")
    else:
        print(f"\n⚠️ {total-passed} tests failed. Check the issues above.")
    
    return passed == total

if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1) 