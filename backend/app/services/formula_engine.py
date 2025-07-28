import re
import ast
import math
import statistics
from typing import Dict, List, Any, Optional, Union, Callable, Set, Tuple
from dataclasses import dataclass
from datetime import datetime, date
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.formula.tokenizer import Tokenizer, Token
from openpyxl.utils import column_index_from_string, get_column_letter
from collections import defaultdict, deque

from app.models.parameter import FormulaNode


def safe_eval(expression: str, context: Dict[str, Any] = None) -> Any:
    """Module-level safe_eval function for backward compatibility."""
    engine = FormulaEngine()
    if context:
        engine.cell_values.update(context)
    return engine._safe_eval(expression)


@dataclass
class CalculationResult:
    """Result of a formula calculation."""
    
    value: Any
    error: Optional[str]
    data_type: str
    calculation_time: float
    dependencies_used: List[str]


@dataclass
class DependencyGraph:
    """Represents formula dependencies."""
    
    nodes: Dict[str, Set[str]]  # cell -> dependencies
    reverse_nodes: Dict[str, Set[str]]  # cell -> dependents
    calculation_order: List[str]
    circular_references: List[List[str]]


class ExcelFunction:
    """Base class for Excel function implementations."""
    
    @staticmethod
    def SUM(*args) -> float:
        """Excel SUM function."""
        total = 0
        for arg in args:
            if isinstance(arg, (list, tuple)):
                total += ExcelFunction.SUM(*arg)
            elif isinstance(arg, (int, float)):
                total += arg
        return total
    
    @staticmethod
    def AVERAGE(*args) -> float:
        """Excel AVERAGE function."""
        values = []
        for arg in args:
            if isinstance(arg, (list, tuple)):
                values.extend(ExcelFunction._flatten_to_numbers(arg))
            elif isinstance(arg, (int, float)):
                values.append(arg)
        
        return sum(values) / len(values) if values else 0
    
    @staticmethod
    def MAX(*args) -> float:
        """Excel MAX function."""
        values = []
        for arg in args:
            if isinstance(arg, (list, tuple)):
                values.extend(ExcelFunction._flatten_to_numbers(arg))
            elif isinstance(arg, (int, float)):
                values.append(arg)
        
        return max(values) if values else 0
    
    @staticmethod
    def MIN(*args) -> float:
        """Excel MIN function."""
        values = []
        for arg in args:
            if isinstance(arg, (list, tuple)):
                values.extend(ExcelFunction._flatten_to_numbers(arg))
            elif isinstance(arg, (int, float)):
                values.append(arg)
        
        return min(values) if values else 0
    
    @staticmethod
    def COUNT(*args) -> int:
        """Excel COUNT function."""
        count = 0
        for arg in args:
            if isinstance(arg, (list, tuple)):
                count += ExcelFunction.COUNT(*arg)
            elif isinstance(arg, (int, float)):
                count += 1
        return count
    
    @staticmethod
    def IF(condition, value_if_true, value_if_false=0):
        """Excel IF function."""
        return value_if_true if condition else value_if_false
    
    @staticmethod
    def ROUND(number, num_digits=0):
        """Excel ROUND function."""
        return round(float(number), int(num_digits))
    
    @staticmethod
    def ABS(number):
        """Excel ABS function."""
        return abs(float(number))
    
    @staticmethod
    def SQRT(number):
        """Excel SQRT function."""
        return math.sqrt(float(number))
    
    @staticmethod
    def POWER(number, power):
        """Excel POWER function."""
        return math.pow(float(number), float(power))
    
    @staticmethod
    def EXP(number):
        """Excel EXP function."""
        return math.exp(float(number))
    
    @staticmethod
    def LN(number):
        """Excel LN function."""
        return math.log(float(number))
    
    @staticmethod
    def LOG(number, base=10):
        """Excel LOG function."""
        number = float(number)
        base = float(base)
        if base == 10:
            return math.log10(number)
        return math.log(number, base)
    
    @staticmethod
    def NPV(rate, *cash_flows):
        """Excel NPV function."""
        npv = 0
        for i, cf in enumerate(cash_flows):
            if isinstance(cf, (list, tuple)):
                for j, flow in enumerate(cf):
                    npv += float(flow) / ((1 + float(rate)) ** (i + j))
            else:
                npv += float(cf) / ((1 + float(rate)) ** i)
        return npv
    
    @staticmethod
    def IRR(cash_flows, guess=0.1):
        """Excel IRR function (simplified)."""
        # Simplified IRR calculation using Newton's method
        rate = guess
        for _ in range(100):  # Max iterations
            npv = 0
            npv_derivative = 0
            for i, cf in enumerate(cash_flows):
                npv += cf / ((1 + rate) ** i)
                if i > 0:
                    npv_derivative -= i * cf / ((1 + rate) ** (i + 1))
            
            if abs(npv) < 1e-10:
                return rate
            
            if npv_derivative == 0:
                break
                
            rate = rate - npv / npv_derivative
        
        return rate
    
    @staticmethod
    def VLOOKUP(lookup_value, table_array, col_index_num, range_lookup=False):
        """Excel VLOOKUP function (simplified)."""
        # This is a simplified implementation
        # In a full implementation, you'd handle the table_array properly
        return lookup_value  # Placeholder
    
    @staticmethod
    def _flatten_to_numbers(data) -> List[float]:
        """Helper to flatten nested data to numbers."""
        result = []
        for item in data:
            if isinstance(item, (list, tuple)):
                result.extend(ExcelFunction._flatten_to_numbers(item))
            elif isinstance(item, (int, float)):
                result.append(float(item))
        return result


class FormulaEngine:
    """Excel formula parsing and calculation engine."""
    
    def __init__(self):
        self.cell_values: Dict[str, Any] = {}
        self.formulas: Dict[str, str] = {}
        self.dependency_graph: Optional[DependencyGraph] = None
        self.excel_functions = ExcelFunction()
        
        # Initialize built-in functions
        self.functions = {
            'SUM': ExcelFunction.SUM,
            'AVERAGE': ExcelFunction.AVERAGE,
            'MAX': ExcelFunction.MAX,
            'MIN': ExcelFunction.MIN,
            'COUNT': ExcelFunction.COUNT,
            'IF': ExcelFunction.IF,
            'ROUND': ExcelFunction.ROUND,
            'ABS': ExcelFunction.ABS,
            'SQRT': ExcelFunction.SQRT,
            'POWER': ExcelFunction.POWER,
            'EXP': ExcelFunction.EXP,
            'LN': ExcelFunction.LN,
            'LOG': ExcelFunction.LOG,
            'NPV': ExcelFunction.NPV,
            'IRR': ExcelFunction.IRR,
            'VLOOKUP': ExcelFunction.VLOOKUP,
        }

    def load_workbook_data(self, workbook_path: str) -> None:
        """
        Load workbook data for formula calculation.
        """
        workbook = openpyxl.load_workbook(workbook_path, data_only=False)
        data_workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            formula_sheet = workbook[sheet_name]
            data_sheet = data_workbook[sheet_name]
            
            for row in formula_sheet.iter_rows():
                for cell in row:
                    cell_ref = f"{sheet_name}!{cell.coordinate}"
                    
                    # Store formula if it exists
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        self.formulas[cell_ref] = cell.value
                    
                    # Store current value
                    data_cell = data_sheet[cell.coordinate]
                    if data_cell.value is not None:
                        self.cell_values[cell_ref] = data_cell.value

    def build_dependency_graph(self) -> DependencyGraph:
        """
        Build dependency graph from formulas.
        """
        nodes = defaultdict(set)
        reverse_nodes = defaultdict(set)
        
        for cell_ref, formula in self.formulas.items():
            dependencies = self._extract_cell_references(formula, cell_ref)
            
            for dep in dependencies:
                nodes[cell_ref].add(dep)
                reverse_nodes[dep].add(cell_ref)
        
        # Calculate calculation order and detect circular references
        calculation_order, circular_refs = self._topological_sort(nodes)
        
        self.dependency_graph = DependencyGraph(
            nodes=dict(nodes),
            reverse_nodes=dict(reverse_nodes),
            calculation_order=calculation_order,
            circular_references=circular_refs
        )
        
        return self.dependency_graph

    def evaluate(self, formula: str, context: Dict[str, Any] = None) -> Any:
        """Evaluate a formula expression for backward compatibility."""
        if context:
            # Update cell values with provided context
            self.cell_values.update(context)
        
        # Security validation - check for dangerous patterns
        if self._is_dangerous_expression(formula):
            raise Exception(
                "Security validation failed: Dangerous expression detected")
        
        # If it's a simple cell reference, return its value
        if re.match(r'^[A-Z]+\d+$', formula):
            return self.cell_values.get(formula, 0)
        
        # If it starts with =, evaluate as formula
        if formula.startswith('='):
            return self._evaluate_formula(formula, 'TEMP!A1')
        
        # Try to evaluate as expression
        return self._safe_eval(formula)

    def calculate_cell(self, cell_ref: str) -> CalculationResult:
        """
        Calculate the value of a specific cell.
        """
        start_time = datetime.utcnow()
        
        try:
            if cell_ref not in self.formulas:
                # Not a formula, return stored value
                value = self.cell_values.get(cell_ref, 0)
                return CalculationResult(
                    value=value,
                    error=None,
                    data_type=type(value).__name__,
                    calculation_time=0,
                    dependencies_used=[]
                )
            
            formula = self.formulas[cell_ref]
            dependencies = self._extract_cell_references(formula, cell_ref)
            
            # Calculate dependencies first
            for dep in dependencies:
                if dep in self.formulas and dep not in self.cell_values:
                    dep_result = self.calculate_cell(dep)
                    if dep_result.error:
                        return CalculationResult(
                            value=None,
                            error=f"Dependency error in {dep}: {dep_result.error}",
                            data_type="error",
                            calculation_time=(datetime.utcnow() - start_time).total_seconds(),
                            dependencies_used=dependencies
                        )
                    self.cell_values[dep] = dep_result.value
            
            # Calculate this cell
            result = self._evaluate_formula(formula, cell_ref)
            
            # Store result
            self.cell_values[cell_ref] = result
            
            return CalculationResult(
                value=result,
                error=None,
                data_type=type(result).__name__,
                calculation_time=(datetime.utcnow() - start_time).total_seconds(),
                dependencies_used=list(dependencies)
            )
            
        except Exception as e:
            return CalculationResult(
                value=None,
                error=str(e),
                data_type="error",
                calculation_time=(datetime.utcnow() - start_time).total_seconds(),
                dependencies_used=[]
            )

    def recalculate_affected_cells(self, changed_cell: str) -> Dict[str, CalculationResult]:
        """
        Recalculate all cells affected by a change.
        """
        if not self.dependency_graph:
            self.build_dependency_graph()
        
        # Get all cells that depend on the changed cell
        affected_cells = self._get_affected_cells(changed_cell)
        
        # Recalculate in dependency order
        results = {}
        
        for cell in affected_cells:
            result = self.calculate_cell(cell)
            results[cell] = result
            
            # If there's an error, stop cascading calculations
            if result.error:
                break
        
        return results

    def _extract_cell_references(self, formula: str, current_cell: str) -> Set[str]:
        """
        Extract cell references from a formula.
        """
        references = set()
        
        try:
            # Parse formula with openpyxl tokenizer
            tokens = Tokenizer(formula).items
            sheet_name = current_cell.split('!')[0] if '!' in current_cell else 'Sheet1'
            
            for token in tokens:
                if token.type == Token.OPERAND and token.subtype == Token.RANGE:
                    ref = token.value
                    
                    # Handle different reference formats
                    if ':' in ref:
                        # Range reference (e.g., A1:B5)
                        start_ref, end_ref = ref.split(':')
                        cell_refs = self._expand_range(start_ref, end_ref, sheet_name)
                        references.update(cell_refs)
                    else:
                        # Single cell reference
                        if '!' not in ref:
                            ref = ref
                        references.add(ref)
        
        except Exception:
            # Fallback to regex parsing
            cell_pattern = r'[A-Z]+\d+'
            matches = re.findall(cell_pattern, formula.upper())
            sheet_name = current_cell.split('!')[0] if '!' in current_cell else 'Sheet1'
            
            for match in matches:
                references.add(match)
        
        return references

    def _expand_range(self, start_ref: str, end_ref: str, sheet_name: str) -> List[str]:
        """
        Expand a cell range to individual cell references.
        """
        # Parse start and end references
        start_col, start_row = self._parse_cell_reference(start_ref)
        end_col, end_row = self._parse_cell_reference(end_ref)
        
        references = []
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                references.append(cell_ref)
        
        return references

    def _parse_cell_reference(self, ref: str) -> Tuple[int, int]:
        """
        Parse cell reference to column and row numbers.
        """
        # Extract column letters and row number
        col_match = re.match(r'([A-Z]+)(\d+)', ref.upper())
        if col_match:
            col_letters, row_num = col_match.groups()
            col_num = column_index_from_string(col_letters)
            return col_num, int(row_num)
        
        raise ValueError(f"Invalid cell reference: {ref}")

    def _topological_sort(self, nodes: Dict[str, Set[str]]) -> Tuple[List[str], List[List[str]]]:
        """
        Perform topological sort to determine calculation order and detect cycles.
        """
        # Kahn's algorithm for topological sorting
        in_degree = defaultdict(int)
        all_nodes = set()
        
        # Calculate in-degrees
        for node, deps in nodes.items():
            all_nodes.add(node)
            all_nodes.update(deps)
            for dep in deps:
                in_degree[dep] += 1
        
        # Initialize queue with nodes having no dependencies
        queue = deque([node for node in all_nodes if in_degree[node] == 0])
        calculation_order = []
        
        while queue:
            current = queue.popleft()
            calculation_order.append(current)
            
            # Process dependents
            for dependent in nodes.get(current, []):
                in_degree[dependent] -= 1
                if in_degree[dependent] == 0:
                    queue.append(dependent)
        
        # Detect circular references
        circular_refs = []
        remaining_nodes = [node for node in all_nodes if in_degree[node] > 0]
        
        if remaining_nodes:
            # Find strongly connected components for circular references
            circular_refs = self._find_circular_references(nodes, remaining_nodes)
        
        return calculation_order, circular_refs

    def _find_circular_references(
        self, 
        nodes: Dict[str, Set[str]], 
        remaining_nodes: List[str]
    ) -> List[List[str]]:
        """
        Find circular reference chains.
        """
        # Simplified circular reference detection
        # In a full implementation, you'd use Tarjan's algorithm
        visited = set()
        circular_refs = []
        
        for node in remaining_nodes:
            if node not in visited:
                path = []
                cycle = self._dfs_cycle_detection(node, nodes, visited, path, set())
                if cycle:
                    circular_refs.append(cycle)
        
        return circular_refs

    def _dfs_cycle_detection(
        self, 
        node: str, 
        nodes: Dict[str, Set[str]], 
        visited: set, 
        path: List[str], 
        path_set: set
    ) -> Optional[List[str]]:
        """
        DFS-based cycle detection.
        """
        if node in path_set:
            # Found a cycle
            cycle_start = path.index(node)
            return path[cycle_start:] + [node]
        
        if node in visited:
            return None
        
        visited.add(node)
        path.append(node)
        path_set.add(node)
        
        for neighbor in nodes.get(node, []):
            cycle = self._dfs_cycle_detection(neighbor, nodes, visited, path, path_set)
            if cycle:
                return cycle
        
        path.pop()
        path_set.remove(node)
        return None

    def _get_affected_cells(self, changed_cell: str) -> List[str]:
        """
        Get all cells affected by a change in topological order.
        """
        if not self.dependency_graph:
            return []
        
        affected = set()
        queue = deque([changed_cell])
        
        while queue:
            current = queue.popleft()
            dependents = self.dependency_graph.reverse_nodes.get(current, set())
            
            for dependent in dependents:
                if dependent not in affected:
                    affected.add(dependent)
                    queue.append(dependent)
        
        # Return in calculation order
        ordered_affected = []
        for cell in self.dependency_graph.calculation_order:
            if cell in affected:
                ordered_affected.append(cell)
        
        return ordered_affected

    def _evaluate_formula(self, formula: str, cell_ref: str) -> Any:
        """
        Evaluate an Excel formula.
        """
        # Remove the leading '='
        if formula.startswith('='):
            formula = formula[1:]
        
        # Replace cell references with values
        processed_formula = self._replace_cell_references(formula, cell_ref)
        
        # Replace Excel functions with Python equivalents
        processed_formula = self._replace_excel_functions(processed_formula)
        
        # Evaluate the expression
        try:
            # Use a restricted eval for safety
            result = self._safe_eval(processed_formula)
            return result
        except Exception as e:
            raise Exception(f"Formula evaluation error: {str(e)}")

    def _replace_cell_references(self, formula: str, current_cell: str) -> str:
        """
        Replace cell references with their values.
        """
        sheet_name = current_cell.split('!')[0] if '!' in current_cell else 'Sheet1'
        
        # Pattern to match ranges like A1:B5
        range_pattern = r'([A-Z]+\d+):([A-Z]+\d+)'
        cell_pattern = r'\b([A-Z]+\d+)\b'

        def replace_range(match):
            start_ref, end_ref = match.groups()
            cells = self._expand_range(start_ref, end_ref, sheet_name)
            values = [self.cell_values.get(c, self.cell_values.get(f"{sheet_name}!{c}", 0)) for c in cells]
            return str(values)

        formula = re.sub(range_pattern, replace_range, formula)
        
        def replace_ref(match):
            ref = match.group(1)
            full_ref = f"{sheet_name}!{ref}"

            value = self.cell_values.get(ref, self.cell_values.get(full_ref, 0))
            
            # Handle different data types
            if isinstance(value, str):
                return f'"{value}"'
            elif value is None:
                return "0"
            else:
                return str(value)
        
        return re.sub(cell_pattern, replace_ref, formula)

    def _replace_excel_functions(self, formula: str) -> str:
        """
        Replace Excel functions with Python function calls.
        """
        # Pattern to match function calls
        func_pattern = r'\b([A-Z_]+)\s*\('
        
        def replace_func(match):
            func_name = match.group(1)
            if func_name in self.functions:
                return f"self.functions['{func_name}']("
            else:
                # Unknown function, replace with a placeholder
                return f"self._unknown_function('{func_name}', "
        
        return re.sub(func_pattern, replace_func, formula)

    def _safe_eval(self, expression: str) -> Any:
        """
        Safely evaluate a mathematical expression.
        """
        # Create a safe environment for evaluation
        safe_dict = {
            'abs': abs,
            'round': round,
            'min': min,
            'max': max,
            'sum': sum,
            'pow': pow,
            '__builtins__': {},
        }

        # Add cell values and variables to the evaluation context
        safe_dict.update(self.cell_values)
        safe_dict['self'] = self
        
        try:
            result = eval(expression, safe_dict)
            return result
        except ZeroDivisionError:
            # Re-raise division by zero to be handled by caller
            raise ZeroDivisionError("Division by zero in formula evaluation")
        except Exception as e:
            raise Exception(f"Expression evaluation failed: {str(e)}")

    def _unknown_function(self, func_name: str, *args) -> Any:
        """
        Handle unknown Excel functions.
        """
        # Unknown function should raise an error for strict evaluation
        print(f"Warning: Unknown function {func_name} called with args {args}")
        raise Exception(f"Unknown function {func_name}")

    def update_cell_value(self, cell_ref: str, new_value: Any) -> Dict[str, CalculationResult]:
        """
        Update a cell value and recalculate affected cells.
        """
        # Store the new value
        self.cell_values[cell_ref] = new_value
        
        # Recalculate affected cells
        return self.recalculate_affected_cells(cell_ref)
    
    def set_cell_value(self, cell_ref: str, new_value: Any) -> Dict[str, CalculationResult]:
        """
        Alias for update_cell_value for backward compatibility with tests.
        """
        return self.update_cell_value(cell_ref, new_value)

    # ------------------------------------------------------------------
    # Convenience helpers used in unit tests
    # ------------------------------------------------------------------

    def get_cell_value(self, cell_ref: str) -> Any:
        """Return stored value for a cell reference."""
        return self.cell_values.get(cell_ref)

    def get_formula(self, cell_ref: str) -> Optional[str]:
        """Return stored formula for a cell reference."""
        return self.formulas.get(cell_ref)

    def set_formula(self, cell_ref: str, formula: str) -> None:
        """Store a formula for later evaluation."""
        self.formulas[cell_ref] = formula

    def evaluate_formula(self, formula: str) -> Any:
        """Public wrapper for evaluating a formula string."""
        return self.evaluate(formula)

    def evaluate_batch(self, formulas: Dict[str, str]) -> Dict[str, Any]:
        """Evaluate a batch of formulas and return a mapping of results."""
        results = {}
        for cell, formula in formulas.items():
            self.set_formula(cell, formula)
            value = self.evaluate_formula(formula)
            results[cell] = value
            self.set_cell_value(cell, value)
        return results

    def get_dependencies(self, cell_ref: str) -> List[str]:
        """Return list of cell references this cell depends on."""
        formula = self.formulas.get(cell_ref)
        if not formula:
            return []
        return list(self._extract_cell_references(formula, cell_ref))

    async def calculate_scenario(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Dummy scenario calculation used in tests."""
        # In real implementation this would trigger workbook evaluation.
        return {"result": sum(parameters.values()) if parameters else 0}

    def get_formula_dependencies(self, cell_ref: str) -> Dict[str, Any]:
        """
        Get detailed dependency information for a cell.
        """
        if not self.dependency_graph:
            self.build_dependency_graph()
        
        return {
            'dependencies': list(self.dependency_graph.nodes.get(cell_ref, set())),
            'dependents': list(self.dependency_graph.reverse_nodes.get(cell_ref, set())),
            'has_formula': cell_ref in self.formulas,
            'formula': self.formulas.get(cell_ref),
            'current_value': self.cell_values.get(cell_ref),
            'calculation_order': self.dependency_graph.calculation_order.index(cell_ref) 
                                if cell_ref in self.dependency_graph.calculation_order else -1
        }

    def detect_circular_references(self) -> List[List[str]]:
        """
        Detect and return circular references in the workbook.
        """
        if not self.dependency_graph:
            self.build_dependency_graph()

        visited = set()
        stack = []

        def dfs(node):
            if node in stack:
                return stack[stack.index(node):] + [node]
            if node in visited:
                return None
            visited.add(node)
            stack.append(node)
            for dep in self.dependency_graph.nodes.get(node, []):
                cycle = dfs(dep)
                if cycle:
                    return cycle
            stack.pop()
            return None

        for cell in self.dependency_graph.nodes:
            cycle = dfs(cell)
            if cycle:
                self.dependency_graph.circular_references.append(cycle)
                raise Exception("Circular reference detected")

        return []

    def get_calculation_statistics(self) -> Dict[str, Any]:
        """
        Get statistics about the calculation engine.
        """
        if not self.dependency_graph:
            self.build_dependency_graph()
        
        return {
            'total_cells': len(self.cell_values),
            'formula_cells': len(self.formulas),
            'dependency_count': sum(len(deps) for deps in self.dependency_graph.nodes.values()),
            'circular_references': len(self.dependency_graph.circular_references),
            'calculation_order_length': len(self.dependency_graph.calculation_order),
            'max_dependency_depth': self._calculate_max_dependency_depth()
        }

    def _calculate_max_dependency_depth(self) -> int:
        """
        Calculate the maximum dependency depth in the formula graph.
        """
        if not self.dependency_graph:
            return 0
        
        depths = {}
        
        def calculate_depth(cell):
            if cell in depths:
                return depths[cell]
            
            deps = self.dependency_graph.nodes.get(cell, set())
            if not deps:
                depths[cell] = 0
                return 0
            
            max_dep_depth = max(calculate_depth(dep) for dep in deps)
            depths[cell] = max_dep_depth + 1
            return depths[cell]
        
        return max(calculate_depth(cell) for cell in self.dependency_graph.nodes.keys()) if self.dependency_graph.nodes else 0 

    def _is_dangerous_expression(self, expression: str) -> bool:
        """
        Check if an expression contains dangerous patterns.
        
        Args:
            expression: The expression to validate
            
        Returns:
            True if the expression is dangerous, False otherwise
        """
        dangerous_patterns = [
            '__import__',
            'exec',
            'eval',
            'open',
            'subprocess',
            'system'
        ]
        
        expression_lower = expression.lower()
        
        # Check for dangerous patterns as whole words/tokens
        import re
        for pattern in dangerous_patterns:
            # Use word boundaries to match complete words only
            if re.search(r'\b' + re.escape(pattern) + r'\b', expression_lower):
                return True
        
        # Check for suspicious characters/patterns
        suspicious_chars = ['__', '{{', '}}', 'system', 'shell', 'subprocess']
        for char in suspicious_chars:
            if char in expression_lower:
                return True
                
        return False 