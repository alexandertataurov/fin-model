import json
import traceback
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass
from enum import Enum
import pandas as pd
from openpyxl import load_workbook

from app.schemas.file import FileValidationResult, ParsedFileData
from app.services.excel_parser import ExcelParser
from app.services.financial_extractor import FinancialExtractor


class ProcessingIssueType(str, Enum):
    """Types of processing issues that can be handled partially."""

    MISSING_DATA = "missing_data"
    INVALID_FORMAT = "invalid_format"
    FORMULA_ERROR = "formula_error"
    SHEET_ACCESS_ERROR = "sheet_access_error"
    VALIDATION_FAILURE = "validation_failure"
    PARSING_ERROR = "parsing_error"
    CORRUPT_DATA = "corrupt_data"


class ProcessingSeverity(str, Enum):
    """Severity levels for processing issues."""

    LOW = "low"  # Can process normally with warnings
    MEDIUM = "medium"  # Can process partially
    HIGH = "high"  # Significant issues but some data recoverable
    CRITICAL = "critical"  # Cannot process meaningfully


@dataclass
class ProcessingIssue:
    """Represents an issue encountered during processing."""

    issue_type: ProcessingIssueType
    severity: ProcessingSeverity
    location: str  # Sheet, cell, or section where issue occurred
    description: str
    suggested_fix: Optional[str] = None
    recoverable: bool = True
    auto_fix_attempted: bool = False
    auto_fix_successful: bool = False


@dataclass
class PartialProcessingResult:
    """Result of partial processing attempt."""

    success: bool
    completion_percentage: float  # 0-100
    extracted_data: Optional[Dict[str, Any]]
    issues_encountered: List[ProcessingIssue]
    sheets_processed: List[str]
    sheets_failed: List[str]
    recovery_actions_taken: List[str]
    recommendations: List[str]


class PartialProcessor:
    """Service for handling partial processing of files with issues."""

    def __init__(self):
        self.excel_parser = ExcelParser()
        self.financial_extractor = FinancialExtractor()
        self.auto_fix_strategies = self._initialize_auto_fix_strategies()

    def process_with_issues(
        self, file_path: str, validation_result: FileValidationResult
    ) -> PartialProcessingResult:
        """
        Attempt to process a file that has validation issues.

        Args:
            file_path: Path to the Excel file
            validation_result: Initial validation result showing issues

        Returns:
            PartialProcessingResult with extracted data and issue details
        """
        issues = []
        extracted_data = {
            "partial_financial_statements": {},
            "recoverable_metrics": {},
            "partial_time_series": {},
            "warnings": [],
        }
        sheets_processed = []
        sheets_failed = []
        recovery_actions = []

        try:
            workbook = load_workbook(file_path, data_only=False)
            total_sheets = len(workbook.sheetnames)

            # Process each sheet individually to maximize recovery
            for sheet_name in workbook.sheetnames:
                try:
                    sheet_result = self._process_sheet_with_recovery(
                        workbook, sheet_name, issues
                    )

                    if sheet_result["success"]:
                        sheets_processed.append(sheet_name)
                        extracted_data["partial_financial_statements"][
                            sheet_name
                        ] = sheet_result["data"]
                        recovery_actions.extend(
                            sheet_result["recovery_actions"]
                        )
                    else:
                        sheets_failed.append(sheet_name)
                        issues.extend(sheet_result["issues"])

                except Exception as e:
                    sheets_failed.append(sheet_name)
                    issues.append(
                        ProcessingIssue(
                            issue_type=ProcessingIssueType.SHEET_ACCESS_ERROR,
                            severity=ProcessingSeverity.HIGH,
                            location=f"Sheet: {sheet_name}",
                            description=f"Failed to access sheet: {str(e)}",
                            recoverable=False,
                        )
                    )

            # Attempt to extract metrics from successfully processed sheets
            if sheets_processed:
                metrics_result = self._extract_partial_metrics(
                    workbook, sheets_processed, issues
                )
                extracted_data["recoverable_metrics"] = metrics_result[
                    "metrics"
                ]
                recovery_actions.extend(metrics_result["recovery_actions"])

            # Calculate completion percentage
            completion_percentage = (
                (len(sheets_processed) / total_sheets) * 100
                if total_sheets > 0
                else 0
            )

            # Adjust completion based on data quality
            if extracted_data["recoverable_metrics"]:
                completion_percentage = min(
                    100, completion_percentage + 20
                )  # Bonus for metrics

            # Generate recommendations
            recommendations = self._generate_recovery_recommendations(
                issues, sheets_failed
            )

            return PartialProcessingResult(
                success=len(sheets_processed) > 0,
                completion_percentage=completion_percentage,
                extracted_data=extracted_data,
                issues_encountered=issues,
                sheets_processed=sheets_processed,
                sheets_failed=sheets_failed,
                recovery_actions_taken=recovery_actions,
                recommendations=recommendations,
            )

        except Exception as e:
            # Complete processing failure
            issues.append(
                ProcessingIssue(
                    issue_type=ProcessingIssueType.PARSING_ERROR,
                    severity=ProcessingSeverity.CRITICAL,
                    location="File level",
                    description=f"Complete processing failure: {str(e)}",
                    recoverable=False,
                )
            )

            return PartialProcessingResult(
                success=False,
                completion_percentage=0,
                extracted_data=None,
                issues_encountered=issues,
                sheets_processed=[],
                sheets_failed=workbook.sheetnames
                if "workbook" in locals()
                else [],
                recovery_actions_taken=[],
                recommendations=[
                    "File may be corrupted",
                    "Try re-uploading the file",
                    "Contact support if issue persists",
                ],
            )

    def _process_sheet_with_recovery(
        self,
        workbook,
        sheet_name: str,
        global_issues: List[ProcessingIssue],
    ) -> Dict[str, Any]:
        """Process a single sheet with error recovery."""
        result = {
            "success": False,
            "data": {},
            "issues": [],
            "recovery_actions": [],
        }

        try:
            sheet = workbook[sheet_name]

            # Basic sheet analysis
            sheet_data = {
                "name": sheet_name,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "extracted_values": {},
                "formulas_found": [],
                "data_ranges": {},
            }

            # Attempt to extract data with recovery
            try:
                # Extract numeric data
                numeric_data = self._extract_numeric_data_with_recovery(
                    sheet, result["issues"]
                )
                sheet_data["extracted_values"] = numeric_data["values"]
                result["recovery_actions"].extend(
                    numeric_data["recovery_actions"]
                )

                # Extract text data and labels
                text_data = self._extract_text_data_with_recovery(
                    sheet, result["issues"]
                )
                sheet_data["labels"] = text_data["labels"]
                result["recovery_actions"].extend(
                    text_data["recovery_actions"]
                )

                # Extract formulas (if accessible)
                formula_data = self._extract_formulas_with_recovery(
                    sheet, result["issues"]
                )
                sheet_data["formulas_found"] = formula_data["formulas"]
                result["recovery_actions"].extend(
                    formula_data["recovery_actions"]
                )

                result["data"] = sheet_data
                result["success"] = True

            except Exception as e:
                result["issues"].append(
                    ProcessingIssue(
                        issue_type=ProcessingIssueType.PARSING_ERROR,
                        severity=ProcessingSeverity.MEDIUM,
                        location=f"Sheet: {sheet_name}",
                        description=f"Partial data extraction failed: {str(e)}",
                        recoverable=True,
                        suggested_fix="Try extracting individual cell ranges",
                    )
                )

                # Attempt basic cell-by-cell recovery
                basic_data = self._basic_cell_recovery(
                    sheet, result["issues"]
                )
                if basic_data:
                    result["data"] = basic_data
                    result["success"] = True
                    result["recovery_actions"].append(
                        "Used basic cell-by-cell recovery"
                    )

        except Exception as e:
            result["issues"].append(
                ProcessingIssue(
                    issue_type=ProcessingIssueType.SHEET_ACCESS_ERROR,
                    severity=ProcessingSeverity.HIGH,
                    location=f"Sheet: {sheet_name}",
                    description=f"Cannot access sheet: {str(e)}",
                    recoverable=False,
                )
            )

        return result

    def _extract_numeric_data_with_recovery(
        self, sheet, issues: List[ProcessingIssue]
    ) -> Dict[str, Any]:
        """Extract numeric data with error recovery."""
        result = {"values": {}, "recovery_actions": []}

        numeric_values = {}
        recovery_count = 0

        for row in sheet.iter_rows(values_only=True):
            for col_idx, cell_value in enumerate(row):
                if isinstance(cell_value, (int, float)):
                    col_letter = chr(65 + col_idx)  # A, B, C, etc.
                    numeric_values[
                        f"{col_letter}{sheet.max_row}"
                    ] = cell_value
                elif isinstance(cell_value, str):
                    # Try to convert string to number
                    try:
                        if (
                            cell_value.replace(".", "")
                            .replace("-", "")
                            .replace(",", "")
                            .isdigit()
                        ):
                            numeric_val = float(
                                cell_value.replace(",", "")
                            )
                            col_letter = chr(65 + col_idx)
                            numeric_values[
                                f"{col_letter}{sheet.max_row}"
                            ] = numeric_val
                            recovery_count += 1
                    except:
                        continue

        result["values"] = numeric_values

        if recovery_count > 0:
            result["recovery_actions"].append(
                f"Recovered {recovery_count} numeric values from text"
            )

        return result

    def _extract_text_data_with_recovery(
        self, sheet, issues: List[ProcessingIssue]
    ) -> Dict[str, Any]:
        """Extract text data with error recovery."""
        result = {"labels": {}, "recovery_actions": []}

        labels = {}

        # Focus on first few rows and columns for labels
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                try:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and len(cell.value.strip()) > 0
                    ):
                        cell_ref = f"{chr(64 + col_idx)}{row_idx}"
                        labels[cell_ref] = cell.value.strip()
                except Exception as e:
                    continue

        result["labels"] = labels

        if labels:
            result["recovery_actions"].append(
                f"Extracted {len(labels)} text labels"
            )

        return result

    def _extract_formulas_with_recovery(
        self, sheet, issues: List[ProcessingIssue]
    ) -> Dict[str, Any]:
        """Extract formulas with error recovery."""
        result = {"formulas": [], "recovery_actions": []}

        formulas = []

        try:
            # Load workbook without data_only to see formulas
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":  # Formula cell
                        try:
                            formulas.append(
                                {
                                    "cell": f"{cell.column_letter}{cell.row}",
                                    "formula": cell.value,
                                    "calculated_value": cell.value
                                    if hasattr(cell, "value")
                                    else None,
                                }
                            )
                        except Exception as e:
                            issues.append(
                                ProcessingIssue(
                                    issue_type=ProcessingIssueType.FORMULA_ERROR,
                                    severity=ProcessingSeverity.LOW,
                                    location=f"Cell: {cell.column_letter}{cell.row}",
                                    description=f"Formula error: {str(e)}",
                                    recoverable=True,
                                )
                            )
        except Exception as e:
            issues.append(
                ProcessingIssue(
                    issue_type=ProcessingIssueType.FORMULA_ERROR,
                    severity=ProcessingSeverity.MEDIUM,
                    location="Sheet formulas",
                    description=f"Cannot extract formulas: {str(e)}",
                    recoverable=True,
                    suggested_fix="Formulas may be corrupted, but values may still be extractable",
                )
            )

        result["formulas"] = formulas

        if formulas:
            result["recovery_actions"].append(
                f"Extracted {len(formulas)} formulas"
            )

        return result

    def _basic_cell_recovery(
        self, sheet, issues: List[ProcessingIssue]
    ) -> Optional[Dict[str, Any]]:
        """Basic cell-by-cell data recovery as last resort."""
        try:
            basic_data = {
                "name": sheet.title,
                "cell_count": 0,
                "data_cells": {},
                "recovery_method": "basic_cell_scan",
            }

            # Scan a reasonable range to avoid huge files
            max_scan_rows = min(100, sheet.max_row)
            max_scan_cols = min(20, sheet.max_column)

            for row in range(1, max_scan_rows + 1):
                for col in range(1, max_scan_cols + 1):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            cell_ref = f"{cell.column_letter}{cell.row}"
                            basic_data["data_cells"][cell_ref] = {
                                "value": cell.value,
                                "type": type(cell.value).__name__,
                            }
                            basic_data["cell_count"] += 1
                    except:
                        continue

            return basic_data if basic_data["cell_count"] > 0 else None

        except Exception as e:
            issues.append(
                ProcessingIssue(
                    issue_type=ProcessingIssueType.PARSING_ERROR,
                    severity=ProcessingSeverity.CRITICAL,
                    location="Basic recovery",
                    description=f"Basic cell recovery failed: {str(e)}",
                    recoverable=False,
                )
            )
            return None

    def _extract_partial_metrics(
        self,
        workbook,
        processed_sheets: List[str],
        issues: List[ProcessingIssue],
    ) -> Dict[str, Any]:
        """Extract what metrics we can from successfully processed sheets."""
        result = {"metrics": {}, "recovery_actions": []}

        try:
            # Use financial extractor on available sheets
            partial_extraction = (
                self.financial_extractor.extract_comprehensive_data(
                    workbook.path
                )
            )

            if partial_extraction and not partial_extraction.get("error"):
                # Filter metrics to only include data from processed sheets
                filtered_metrics = []

                for metric in partial_extraction.get(
                    "financial_metrics", []
                ):
                    # Check if metric comes from a processed sheet
                    metric_sheet = self._identify_metric_sheet(
                        metric, processed_sheets
                    )
                    if metric_sheet:
                        filtered_metrics.append(metric)

                result["metrics"] = {
                    "financial_metrics": filtered_metrics,
                    "extraction_confidence": len(filtered_metrics)
                    / max(
                        len(
                            partial_extraction.get("financial_metrics", [])
                        ),
                        1,
                    ),
                    "sheets_contributing": processed_sheets,
                }

                result["recovery_actions"].append(
                    f"Extracted {len(filtered_metrics)} financial metrics from {len(processed_sheets)} sheets"
                )

        except Exception as e:
            issues.append(
                ProcessingIssue(
                    issue_type=ProcessingIssueType.PARSING_ERROR,
                    severity=ProcessingSeverity.MEDIUM,
                    location="Metrics extraction",
                    description=f"Partial metrics extraction failed: {str(e)}",
                    recoverable=True,
                    suggested_fix="Manual review of specific metrics may be needed",
                )
            )

        return result

    def _identify_metric_sheet(
        self, metric: Dict[str, Any], processed_sheets: List[str]
    ) -> Optional[str]:
        """Identify which sheet a metric came from."""
        # Check source_cells for sheet references
        source_cells = metric.get("source_cells", [])
        for cell_ref in source_cells:
            for sheet_name in processed_sheets:
                if sheet_name.lower() in cell_ref.lower():
                    return sheet_name

        # If no explicit sheet reference, assume it came from a processed sheet
        return processed_sheets[0] if processed_sheets else None

    def _generate_recovery_recommendations(
        self, issues: List[ProcessingIssue], failed_sheets: List[str]
    ) -> List[str]:
        """Generate recommendations based on processing issues."""
        recommendations = []

        # Analyze issue patterns
        critical_issues = [
            i for i in issues if i.severity == ProcessingSeverity.CRITICAL
        ]
        high_issues = [
            i for i in issues if i.severity == ProcessingSeverity.HIGH
        ]

        if critical_issues:
            recommendations.append(
                "File has critical issues that prevent complete processing"
            )
            recommendations.append(
                "Consider manually fixing the file before re-uploading"
            )

        if failed_sheets:
            recommendations.append(
                f"Could not process {len(failed_sheets)} sheets: {', '.join(failed_sheets[:3])}"
            )
            if len(failed_sheets) > 3:
                recommendations.append(
                    f"...and {len(failed_sheets) - 3} more sheets"
                )

        # Issue-specific recommendations
        issue_types = [i.issue_type for i in issues]

        if ProcessingIssueType.FORMULA_ERROR in issue_types:
            recommendations.append(
                "Some formulas could not be evaluated - check for circular references or missing data"
            )

        if ProcessingIssueType.MISSING_DATA in issue_types:
            recommendations.append(
                "Fill in missing data fields for more complete analysis"
            )

        if ProcessingIssueType.INVALID_FORMAT in issue_types:
            recommendations.append(
                "Review data formatting - some values may not be recognized as numbers"
            )

        # General recommendations
        if not recommendations:
            recommendations.append(
                "Processing completed successfully with partial data extraction"
            )

        recommendations.append(
            "Review the extracted data to ensure it meets your requirements"
        )

        return recommendations

    def _initialize_auto_fix_strategies(self) -> Dict[str, Any]:
        """Initialize automatic fix strategies for common issues."""
        return {
            "text_to_number": self._auto_fix_text_to_number,
            "missing_headers": self._auto_fix_missing_headers,
            "formula_errors": self._auto_fix_formula_errors,
            "date_format": self._auto_fix_date_format,
        }

    def _auto_fix_text_to_number(self, cell_value: str) -> Optional[float]:
        """Auto-fix text that should be numbers."""
        try:
            # Remove common formatting
            cleaned = (
                cell_value.replace(",", "")
                .replace("$", "")
                .replace("%", "")
                .strip()
            )
            if cleaned.replace(".", "").replace("-", "").isdigit():
                return float(cleaned)
        except:
            pass
        return None

    def _auto_fix_missing_headers(self, sheet) -> List[str]:
        """Auto-fix missing or unclear headers."""
        # Generate generic headers if missing
        return [f"Column_{i+1}" for i in range(sheet.max_column)]

    def _auto_fix_formula_errors(self, formula: str) -> str:
        """Attempt to fix common formula errors."""
        # Replace common errors
        fixes = {"#DIV/0!": "0", "#N/A": "0", "#VALUE!": "0", "#REF!": "0"}

        for error, replacement in fixes.items():
            if error in formula:
                return formula.replace(error, replacement)

        return formula

    def _auto_fix_date_format(self, date_value: str) -> Optional[datetime]:
        """Auto-fix date formatting issues."""
        import dateutil.parser

        try:
            return dateutil.parser.parse(date_value)
        except:
            return None
