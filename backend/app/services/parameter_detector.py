import re
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple, Set
from dataclasses import dataclass
from datetime import datetime
import openpyxl
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.formula.translate import Translator

from app.models.parameter import ParameterType, ParameterCategory, SensitivityLevel
from app.services.excel_parser import ExcelParser


@dataclass
class DetectedParameter:
    """Data class for detected parameters."""

    cell_reference: str
    sheet_name: str
    name: str
    value: Optional[float]
    parameter_type: ParameterType
    category: ParameterCategory
    sensitivity_level: SensitivityLevel
    description: Optional[str]
    unit: Optional[str]
    format_type: str
    min_value: Optional[float]
    max_value: Optional[float]
    depends_on: List[str]
    affects: List[str]
    formula: Optional[str]
    validation_rules: Dict[str, Any]
    confidence_score: float


class ParameterDetector:
    """Service for detecting and classifying parameters from Excel files."""

    def __init__(self):
        self.excel_parser = ExcelParser()

        # Parameter detection patterns
        self.growth_patterns = [
            r"growth|rate|cagr|increase|expansion",
            r"%|percent|pct",
        ]

        self.revenue_patterns = [
            r"revenue|sales|income|turnover",
            r"price|pricing|unit.*price",
            r"volume|quantity|units",
            r"market.*share|penetration",
        ]

        self.cost_patterns = [
            r"cost|expense|expenditure|spending",
            r"cogs|cost.*goods.*sold",
            r"opex|operational.*expense",
            r"capex|capital.*expenditure",
            r"depreciation|amortization",
        ]

        self.financial_patterns = [
            r"discount.*rate|wacc|hurdle.*rate",
            r"tax.*rate|effective.*tax",
            r"interest.*rate|borrowing.*rate",
            r"margin|ebitda|profit",
            r"debt|equity|leverage|ratio",
        ]

        self.operational_patterns = [
            r"employees|headcount|staff",
            r"capacity|utilization|efficiency",
            r"inventory|working.*capital",
            r"days.*sales|collection.*period",
        ]

    def detect_growth_patterns(self, data: List[float]) -> Dict[str, Any]:
        """Detect growth rates in a simple time series list."""
        result = {
            "average_growth_rate": 0.0,
            "growth_rates": [],
        }

        try:
            if isinstance(data, list) and len(data) > 1:
                rates = []
                for prev, curr in zip(data[:-1], data[1:]):
                    if prev != 0:
                        rates.append((curr - prev) / prev)
                if rates:
                    result["growth_rates"] = rates
                    result["average_growth_rate"] = sum(rates) / len(rates)
        except Exception as e:
            result["error"] = str(e)

        return result

    async def detect_parameters(
        self, file_path: str, user_id: int
    ) -> List[DetectedParameter]:
        """
        Detect and classify parameters from an Excel file.
        """
        try:
            # Parse the Excel file
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            data_workbook = openpyxl.load_workbook(file_path, data_only=True)

            detected_parameters = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data_sheet = data_workbook[sheet_name]

                # Detect parameters in this sheet
                sheet_parameters = await self._detect_sheet_parameters(
                    sheet, data_sheet, sheet_name
                )
                detected_parameters.extend(sheet_parameters)

            # Build dependency graph
            await self._build_dependency_graph(detected_parameters, workbook)

            # Classify sensitivity levels
            await self._classify_sensitivity_levels(detected_parameters)

            return detected_parameters

        except Exception as e:
            raise Exception(f"Failed to detect parameters: {str(e)}")

    async def _detect_sheet_parameters(
        self, formula_sheet, data_sheet, sheet_name: str
    ) -> List[DetectedParameter]:
        """
        Detect parameters in a specific sheet.
        """
        parameters = []

        # Scan all cells for potential parameters
        for row in formula_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parameter = await self._analyze_cell(cell, data_sheet, sheet_name)
                    if parameter:
                        parameters.append(parameter)

        return parameters

    async def _analyze_cell(
        self, formula_cell, data_cell_sheet, sheet_name: str
    ) -> Optional[DetectedParameter]:
        """
        Analyze a cell to determine if it's a parameter.
        """
        cell_ref = f"{sheet_name}!{formula_cell.coordinate}"
        data_cell = data_cell_sheet[formula_cell.coordinate]

        # Skip if no value
        if formula_cell.value is None and data_cell.value is None:
            return None

        # Check if it's a formula
        is_formula = isinstance(
            formula_cell.value, str
        ) and formula_cell.value.startswith("=")

        # Get the actual value
        actual_value = (
            data_cell.value if data_cell.value is not None else formula_cell.value
        )

        # Try to convert to float
        numeric_value = self._extract_numeric_value(actual_value)

        # Determine if this cell is likely a parameter
        if not await self._is_likely_parameter(formula_cell, data_cell, is_formula):
            return None

        # Classify the parameter
        param_type, category = await self._classify_parameter(
            formula_cell, sheet_name, is_formula
        )

        # Generate parameter name
        param_name = await self._generate_parameter_name(formula_cell, sheet_name)

        # Determine format type
        format_type = self._determine_format_type(actual_value, param_type)

        # Extract validation rules
        validation_rules = await self._extract_validation_rules(
            formula_cell, param_type
        )

        # Calculate confidence score
        confidence_score = await self._calculate_confidence_score(
            formula_cell, param_type, category
        )

        return DetectedParameter(
            cell_reference=cell_ref,
            sheet_name=sheet_name,
            name=param_name,
            value=numeric_value,
            parameter_type=param_type,
            category=category,
            sensitivity_level=SensitivityLevel.MEDIUM,  # Will be updated later
            description=await self._generate_description(formula_cell, param_type),
            unit=self._detect_unit(formula_cell),
            format_type=format_type,
            min_value=None,  # Will be set based on validation rules
            max_value=None,  # Will be set based on validation rules
            depends_on=[],  # Will be populated in dependency analysis
            affects=[],  # Will be populated in dependency analysis
            formula=formula_cell.value if is_formula else None,
            validation_rules=validation_rules,
            confidence_score=confidence_score,
        )

    async def _is_likely_parameter(
        self, formula_cell, data_cell, is_formula: bool
    ) -> bool:
        """
        Determine if a cell is likely to be a parameter.
        """
        # Parameters are typically:
        # 1. Single numeric values (not formulas with complex calculations)
        # 2. Located in assumption areas
        # 3. Have descriptive labels nearby
        # 4. Are referenced by other formulas

        # Check if it's a simple value or simple formula
        if is_formula:
            formula = formula_cell.value.lower()
            # Skip complex formulas (more than 2 operators)
            operators = len(re.findall(r"[+\-*/]", formula))
            if operators > 2:
                return False

            # Skip functions that are clearly calculations (SUM, AVERAGE, etc.)
            calc_functions = ["sum", "average", "max", "min", "count", "if"]
            if any(func in formula for func in calc_functions):
                return False

        # Check if it has a numeric value
        actual_value = (
            data_cell.value if data_cell.value is not None else formula_cell.value
        )
        if not self._has_numeric_component(actual_value):
            return False

        # Check position - parameters are often in the top-left area
        row, col = formula_cell.row, formula_cell.column
        if row > 100 or col > 20:  # Adjust these thresholds as needed
            return False

        # Check for nearby labels
        if await self._has_descriptive_label(formula_cell):
            return True

        # Check if it's referenced by other cells
        if await self._is_referenced_by_formulas(formula_cell):
            return True

        return False

    async def _classify_parameter(
        self, cell, sheet_name: str, is_formula: bool
    ) -> Tuple[ParameterType, ParameterCategory]:
        """
        Classify the parameter type and category.
        """
        # Get context from nearby cells and cell value
        context = await self._get_cell_context(cell, sheet_name)
        context_text = " ".join(context).lower()

        # Classify based on patterns
        param_type = ParameterType.CONSTANT
        category = ParameterCategory.ASSUMPTIONS

        # Check for growth rates
        if any(re.search(pattern, context_text) for pattern in self.growth_patterns):
            if any(re.search(r"%|percent", context_text) for pattern in [context_text]):
                param_type = ParameterType.GROWTH_RATE
                category = ParameterCategory.ASSUMPTIONS

        # Check for revenue parameters
        elif any(re.search(pattern, context_text) for pattern in self.revenue_patterns):
            if "price" in context_text:
                param_type = ParameterType.PRICE
            elif "volume" in context_text or "quantity" in context_text:
                param_type = ParameterType.VOLUME
            else:
                param_type = ParameterType.REVENUE_ASSUMPTION
            category = ParameterCategory.REVENUE

        # Check for cost parameters
        elif any(re.search(pattern, context_text) for pattern in self.cost_patterns):
            param_type = ParameterType.COST_ASSUMPTION
            category = ParameterCategory.COSTS

        # Check for financial parameters
        elif any(
            re.search(pattern, context_text) for pattern in self.financial_patterns
        ):
            if "discount" in context_text or "wacc" in context_text:
                param_type = ParameterType.DISCOUNT_RATE
            elif "tax" in context_text:
                param_type = ParameterType.TAX_RATE
            elif "interest" in context_text:
                param_type = ParameterType.INTEREST_RATE
            else:
                param_type = ParameterType.RATIO_ASSUMPTION
            category = ParameterCategory.FINANCIAL

        # Check for operational parameters
        elif any(
            re.search(pattern, context_text) for pattern in self.operational_patterns
        ):
            param_type = ParameterType.VARIABLE
            category = ParameterCategory.OPERATIONS

        # If it's a formula, it might be calculated
        if is_formula:
            param_type = ParameterType.CALCULATED
            category = ParameterCategory.CALCULATED

        return param_type, category

    async def _get_cell_context(self, cell, sheet_name: str) -> List[str]:
        """
        Get text context from nearby cells to help with classification.
        """
        context = []
        sheet = cell.parent

        # Check cells around this cell (3x3 grid)
        for row_offset in range(-2, 3):
            for col_offset in range(-2, 3):
                try:
                    target_row = cell.row + row_offset
                    target_col = cell.column + col_offset

                    if target_row > 0 and target_col > 0:
                        target_cell = sheet.cell(target_row, target_col)
                        if target_cell.value and isinstance(target_cell.value, str):
                            context.append(target_cell.value)
                except:
                    continue

        return context

    async def _generate_parameter_name(self, cell, sheet_name: str) -> str:
        """
        Generate a meaningful name for the parameter.
        """
        # Try to find a label in nearby cells
        label = await self._find_nearest_label(cell)

        if label:
            # Clean up the label
            name = re.sub(r"[^\w\s]", "", label)
            name = re.sub(r"\s+", "_", name.strip())
            return name.lower()

        # Fallback to sheet and cell reference
        return f"{sheet_name.lower()}_{cell.coordinate.lower()}"

    async def _find_nearest_label(self, cell) -> Optional[str]:
        """
        Find the nearest text label for this cell.
        """
        sheet = cell.parent

        # Check left and above (common label positions)
        candidates = []

        # Check cell to the left
        if cell.column > 1:
            left_cell = sheet.cell(cell.row, cell.column - 1)
            if left_cell.value and isinstance(left_cell.value, str):
                candidates.append((left_cell.value, 1))  # (text, priority)

        # Check cell above
        if cell.row > 1:
            above_cell = sheet.cell(cell.row - 1, cell.column)
            if above_cell.value and isinstance(above_cell.value, str):
                candidates.append((above_cell.value, 2))

        # Check cell above-left (diagonal)
        if cell.row > 1 and cell.column > 1:
            diag_cell = sheet.cell(cell.row - 1, cell.column - 1)
            if diag_cell.value and isinstance(diag_cell.value, str):
                candidates.append((diag_cell.value, 3))

        # Return the highest priority candidate
        if candidates:
            candidates.sort(key=lambda x: x[1])
            return candidates[0][0]

        return None

    def _extract_numeric_value(self, value: Any) -> Optional[float]:
        """
        Extract numeric value from cell content.
        """
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            # Remove common non-numeric characters
            clean_str = re.sub(r"[^\d.\-+e]", "", value)
            try:
                return float(clean_str)
            except:
                return None

        return None

    def _has_numeric_component(self, value: Any) -> bool:
        """
        Check if value has a numeric component.
        """
        return self._extract_numeric_value(value) is not None

    def _determine_format_type(self, value: Any, param_type: ParameterType) -> str:
        """
        Determine the format type for the parameter.
        """
        if param_type in [
            ParameterType.GROWTH_RATE,
            ParameterType.TAX_RATE,
            ParameterType.INTEREST_RATE,
            ParameterType.DISCOUNT_RATE,
        ]:
            return "percentage"

        if param_type in [
            ParameterType.PRICE,
            ParameterType.REVENUE_ASSUMPTION,
            ParameterType.COST_ASSUMPTION,
        ]:
            return "currency"

        return "number"

    def _detect_unit(self, cell) -> Optional[str]:
        """
        Detect the unit of measurement for the parameter.
        """
        # Look for units in nearby cells or cell formatting
        context = []
        sheet = cell.parent

        # Check adjacent cells for unit indicators
        for row_offset in [-1, 0, 1]:
            for col_offset in [-1, 0, 1]:
                try:
                    target_row = cell.row + row_offset
                    target_col = cell.column + col_offset

                    if target_row > 0 and target_col > 0:
                        target_cell = sheet.cell(target_row, target_col)
                        if target_cell.value and isinstance(target_cell.value, str):
                            context.append(target_cell.value.lower())
                except:
                    continue

        context_text = " ".join(context)

        # Common unit patterns
        if re.search(r"%|percent|pct", context_text):
            return "%"
        elif re.search(r"\$|dollar|usd", context_text):
            return "$"
        elif re.search(r"€|euro|eur", context_text):
            return "€"
        elif re.search(r"£|pound|gbp", context_text):
            return "£"
        elif re.search(r"days|day", context_text):
            return "days"
        elif re.search(r"years|year|yr", context_text):
            return "years"
        elif re.search(r"months|month|mo", context_text):
            return "months"

        return None

    async def _extract_validation_rules(
        self, cell, param_type: ParameterType
    ) -> Dict[str, Any]:
        """
        Extract validation rules for the parameter.
        """
        rules = {}

        # Set default ranges based on parameter type
        if param_type in [
            ParameterType.GROWTH_RATE,
            ParameterType.TAX_RATE,
            ParameterType.INTEREST_RATE,
        ]:
            rules["min_value"] = 0.0
            rules["max_value"] = (
                100.0
                if self._determine_format_type(None, param_type) == "percentage"
                else 1.0
            )

        elif param_type == ParameterType.DISCOUNT_RATE:
            rules["min_value"] = 0.0
            rules["max_value"] = 50.0

        elif param_type in [ParameterType.PRICE, ParameterType.VOLUME]:
            rules["min_value"] = 0.0

        # Add data type validation
        rules["data_type"] = "float"
        rules["required"] = True

        return rules

    async def _calculate_confidence_score(
        self, cell, param_type: ParameterType, category: ParameterCategory
    ) -> float:
        """
        Calculate confidence score for parameter detection.
        """
        score = 0.5  # Base score

        # Increase score if we found a good label
        if await self._find_nearest_label(cell):
            score += 0.2

        # Increase score based on parameter type classification
        if param_type != ParameterType.CONSTANT:
            score += 0.2

        # Increase score if it's in a typical parameter location
        if cell.row <= 20 and cell.column <= 10:
            score += 0.1

        return min(score, 1.0)

    async def _has_descriptive_label(self, cell) -> bool:
        """
        Check if cell has a descriptive label nearby.
        """
        return await self._find_nearest_label(cell) is not None

    async def _is_referenced_by_formulas(self, cell) -> bool:
        """
        Check if this cell is referenced by other formulas.
        """
        # This would require scanning the entire workbook
        # For now, return a simple heuristic
        return cell.row <= 50 and cell.column <= 10

    async def _build_dependency_graph(
        self, parameters: List[DetectedParameter], workbook
    ) -> None:
        """
        Build dependency relationships between parameters.
        """
        # Create a mapping of cell references to parameters
        param_map = {param.cell_reference: param for param in parameters}

        for param in parameters:
            if param.formula:
                # Parse formula to find dependencies
                dependencies = self._parse_formula_dependencies(
                    param.formula, param.sheet_name
                )

                for dep_ref in dependencies:
                    if dep_ref in param_map:
                        param.depends_on.append(dep_ref)
                        param_map[dep_ref].affects.append(param.cell_reference)

    def _parse_formula_dependencies(self, formula: str, sheet_name: str) -> List[str]:
        """
        Parse formula to extract cell dependencies.
        """
        dependencies = []

        try:
            # Use openpyxl tokenizer to parse formula
            tokens = Tokenizer(formula).items

            for token in tokens:
                if token.type == "OPERAND" and token.subtype == "RANGE":
                    # This is a cell reference
                    ref = token.value

                    # Add sheet name if not present
                    if "!" not in ref:
                        ref = f"{sheet_name}!{ref}"

                    dependencies.append(ref)

        except Exception:
            # Fallback to regex parsing
            cell_pattern = r"[A-Z]+\d+"
            matches = re.findall(cell_pattern, formula.upper())

            for match in matches:
                ref = f"{sheet_name}!{match}"
                dependencies.append(ref)

        return dependencies

    async def _classify_sensitivity_levels(
        self, parameters: List[DetectedParameter]
    ) -> None:
        """
        Classify sensitivity levels for parameters.
        """
        for param in parameters:
            # Classify based on parameter type and impact
            if param.parameter_type in [
                ParameterType.GROWTH_RATE,
                ParameterType.DISCOUNT_RATE,
                ParameterType.REVENUE_ASSUMPTION,
            ]:
                param.sensitivity_level = SensitivityLevel.HIGH

            elif param.parameter_type in [
                ParameterType.TAX_RATE,
                ParameterType.INTEREST_RATE,
                ParameterType.COST_ASSUMPTION,
            ]:
                param.sensitivity_level = SensitivityLevel.MEDIUM

            elif len(param.affects) > 5:  # High impact parameters
                param.sensitivity_level = SensitivityLevel.HIGH

            elif len(param.affects) > 2:  # Medium impact parameters
                param.sensitivity_level = SensitivityLevel.MEDIUM

            else:
                param.sensitivity_level = SensitivityLevel.LOW

    async def _generate_description(
        self, cell, param_type: ParameterType
    ) -> Optional[str]:
        """
        Generate a description for the parameter.
        """
        # Get context and create a meaningful description
        context = await self._get_cell_context(cell, cell.parent.title)

        if context:
            # Use the first meaningful text as description
            for text in context:
                if len(text) > 5 and not re.match(r"^[A-Z]+\d+$", text):
                    return text

        # Default descriptions based on parameter type
        descriptions = {
            ParameterType.GROWTH_RATE: "Growth rate assumption",
            ParameterType.DISCOUNT_RATE: "Discount rate for valuation",
            ParameterType.TAX_RATE: "Tax rate assumption",
            ParameterType.PRICE: "Unit price assumption",
            ParameterType.VOLUME: "Volume/quantity assumption",
            ParameterType.REVENUE_ASSUMPTION: "Revenue-related assumption",
            ParameterType.COST_ASSUMPTION: "Cost-related assumption",
        }

        return descriptions.get(param_type, "Financial parameter")

    def detect_from_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Simplified detection method for parameter service integration.
        Returns list of detected parameter data for parameter service.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            data_workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            detected_params = []
            
            for sheet_name in workbook.sheetnames:
                formula_sheet = workbook[sheet_name]
                data_sheet = data_workbook[sheet_name]
                
                # Scan first 20 rows and 10 columns for parameters
                for row in range(1, 21):
                    for col in range(1, 11):
                        try:
                            formula_cell = formula_sheet.cell(row, col)
                            data_cell = data_sheet.cell(row, col)
                            
                            if self._is_simple_parameter(formula_cell, data_cell):
                                param_data = self._extract_parameter_data(
                                    formula_cell, data_cell, sheet_name
                                )
                                if param_data:
                                    detected_params.append(param_data)
                        except:
                            continue
            
            return detected_params
            
        except Exception as e:
            # Return empty list if detection fails
            return []
    
    def _is_simple_parameter(self, formula_cell, data_cell) -> bool:
        """Check if cell is a simple parameter (numeric value, not complex formula)."""
        # Must have a numeric value
        value = data_cell.value if data_cell.value is not None else formula_cell.value
        if not self._has_numeric_component(value):
            return False
        
        # If it's a formula, it should be simple
        if isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
            # Skip complex formulas
            formula = formula_cell.value.lower()
            complex_functions = ['sum', 'average', 'vlookup', 'index', 'match', 'if']
            if any(func in formula for func in complex_functions):
                return False
        
        return True
    
    def _extract_parameter_data(self, formula_cell, data_cell, sheet_name: str) -> Dict[str, Any]:
        """Extract parameter data from cell for parameter service."""
        value = data_cell.value if data_cell.value is not None else formula_cell.value
        numeric_value = self._extract_numeric_value(value)
        
        # Get context for classification
        context = self._get_cell_context_sync(formula_cell, sheet_name)
        context_text = ' '.join(context).lower()
        
        # Classify parameter
        param_type, category = self._classify_parameter_sync(context_text, formula_cell)
        
        # Generate name
        name = self._generate_name_sync(formula_cell, context)
        
        return {
            'name': name,
            'value': numeric_value,
            'type': param_type,
            'category': category,
            'sheet': sheet_name,
            'cell': formula_cell.coordinate,
            'min_value': self._infer_min_value(param_type),
            'max_value': self._infer_max_value(param_type)
        }
    
    def _get_cell_context_sync(self, cell, sheet_name: str) -> List[str]:
        """Synchronous version of context extraction."""
        context = []
        sheet = cell.parent
        
        # Check adjacent cells for labels
        for row_offset in [-1, 0]:
            for col_offset in [-1, 0]:
                if row_offset == 0 and col_offset == 0:
                    continue
                try:
                    target_row = cell.row + row_offset
                    target_col = cell.column + col_offset
                    if target_row > 0 and target_col > 0:
                        target_cell = sheet.cell(target_row, target_col)
                        if target_cell.value and isinstance(target_cell.value, str):
                            context.append(target_cell.value)
                except:
                    continue
        
        return context
    
    def _classify_parameter_sync(self, context_text: str, cell) -> Tuple[ParameterType, ParameterCategory]:
        """Synchronous parameter classification."""
        # Check for growth/rate patterns
        if any(pattern in context_text for pattern in ['growth', 'rate', '%', 'percent']):
            return ParameterType.GROWTH_RATE, ParameterCategory.ASSUMPTIONS
        
        # Check for revenue patterns
        if any(pattern in context_text for pattern in ['revenue', 'sales', 'price', 'volume']):
            return ParameterType.REVENUE_ASSUMPTION, ParameterCategory.REVENUE
        
        # Check for cost patterns
        if any(pattern in context_text for pattern in ['cost', 'expense', 'cogs']):
            return ParameterType.COST_ASSUMPTION, ParameterCategory.COSTS
        
        # Check for financial patterns
        if any(pattern in context_text for pattern in ['discount', 'wacc', 'tax', 'interest']):
            return ParameterType.INTEREST_RATE, ParameterCategory.FINANCIAL
        
        # Default
        return ParameterType.CONSTANT, ParameterCategory.ASSUMPTIONS
    
    def _generate_name_sync(self, cell, context: List[str]) -> str:
        """Generate parameter name from context."""
        if context:
            # Use first meaningful text
            for text in context:
                if len(text) > 2 and not text.isdigit():
                    # Clean up the text
                    name = re.sub(r'[^\w\s]', '', text)
                    name = re.sub(r'\s+', '_', name.strip())
                    return name.lower()
        
        # Fallback to cell reference
        return f"param_{cell.coordinate.lower()}"
    
    def _infer_min_value(self, param_type: ParameterType) -> Optional[float]:
        """Infer minimum value based on parameter type."""
        if param_type in [ParameterType.GROWTH_RATE, ParameterType.TAX_RATE, 
                         ParameterType.INTEREST_RATE, ParameterType.DISCOUNT_RATE]:
            return 0.0
        elif param_type in [ParameterType.PRICE, ParameterType.VOLUME]:
            return 0.0
        return None
    
    def _infer_max_value(self, param_type: ParameterType) -> Optional[float]:
        """Infer maximum value based on parameter type."""
        if param_type in [ParameterType.GROWTH_RATE, ParameterType.TAX_RATE]:
            return 1.0  # 100%
        elif param_type == ParameterType.DISCOUNT_RATE:
            return 0.5  # 50%
        return None

    def detect_seasonality(self, data: List[float]) -> Dict[str, Any]:
        """Detect simple seasonal patterns in a numeric series."""
        result = {
            "has_seasonality": False,
            "seasonal_periods": [],
            "seasonal_strength": 0.0,
            "seasonal_pattern": [],
        }

        try:
            if isinstance(data, list) and len(data) >= 8:
                period = 4
                groups = [data[i::period] for i in range(period)]
                averages = [sum(g) / len(g) for g in groups if g]
                overall_avg = sum(data) / len(data)
                if overall_avg != 0:
                    deviations = [
                        abs(a - overall_avg) / abs(overall_avg) for a in averages
                    ]
                    strength = 1 - (sum(deviations) / len(deviations))
                    result["seasonal_strength"] = max(0.0, strength)
                    result["seasonal_pattern"] = averages
                    if strength > 0.2:
                        result["has_seasonality"] = True
                        result["seasonal_periods"] = [period]
        except Exception as e:
            result["error"] = str(e)

        return result

    def identify_assumptions(self, parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Identify key assumptions from parsed financial data.

        Args:
            parsed_data: Parsed financial data from Excel file

        Returns:
            Dictionary containing identified assumptions
        """
        assumptions = {
            "growth_assumptions": [],
            "operational_assumptions": [],
            "financial_assumptions": [],
            "confidence_scores": {},
        }

        try:
            # Extract assumptions from sheets
            if "sheets" in parsed_data:
                for sheet in parsed_data["sheets"]:
                    sheet_name = sheet.get("name", "").lower()

                    # Look for assumption-related sheets
                    if any(
                        keyword in sheet_name
                        for keyword in ["assumption", "input", "parameter", "scenario"]
                    ):
                        # Extract numeric values that look like assumptions
                        if "cells" in sheet:
                            for cell in sheet["cells"]:
                                if self._looks_like_assumption(cell):
                                    assumption = self._create_assumption_from_cell(
                                        cell, sheet_name
                                    )
                                    if assumption:
                                        category = self._categorize_assumption(
                                            assumption
                                        )
                                        assumptions[f"{category}_assumptions"].append(
                                            assumption
                                        )

            # Calculate confidence scores
            for category in ["growth", "operational", "financial"]:
                key = f"{category}_assumptions"
                if assumptions[key]:
                    assumptions["confidence_scores"][category] = min(
                        0.9, len(assumptions[key]) / 10.0
                    )  # More assumptions = higher confidence

        except Exception as e:
            assumptions["error"] = str(e)

        return assumptions

    def validate_assumptions(self, assumptions: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate the reasonableness of financial assumptions.

        Args:
            assumptions: Dictionary of assumptions to validate

        Returns:
            Dictionary containing validation results
        """
        validation_result = {
            "is_valid": True,
            "warnings": [],
            "errors": [],
            "recommendations": [],
        }

        try:
            # Validate growth assumptions
            if "growth_assumptions" in assumptions:
                for assumption in assumptions["growth_assumptions"]:
                    if isinstance(assumption, dict) and "value" in assumption:
                        try:
                            value = float(assumption["value"])

                            # Check for unrealistic growth rates
                            if abs(value) > 1.0:  # More than 100% growth
                                validation_result["warnings"].append(
                                    f"High growth rate detected: {value*100:.1f}%"
                                )
                            elif abs(value) > 0.5:  # More than 50% growth
                                validation_result["warnings"].append(
                                    f"Aggressive growth rate: {value*100:.1f}%"
                                )

                        except (ValueError, TypeError):
                            validation_result["errors"].append(
                                f"Invalid growth rate value: {assumption.get('value')}"
                            )

            # Validate operational assumptions
            if "operational_assumptions" in assumptions:
                for assumption in assumptions["operational_assumptions"]:
                    if isinstance(assumption, dict) and "value" in assumption:
                        try:
                            value = float(assumption["value"])
                            name = assumption.get("name", "").lower()

                            # Check margin assumptions
                            if "margin" in name and (value < 0 or value > 1):
                                validation_result["errors"].append(
                                    f"Invalid margin value: {value}"
                                )

                        except (ValueError, TypeError):
                            continue

            # Add recommendations
            if validation_result["warnings"]:
                validation_result["recommendations"].append(
                    "Review aggressive assumptions for reasonableness"
                )
            if validation_result["errors"]:
                validation_result["recommendations"].append(
                    "Correct invalid assumption values"
                )
                validation_result["is_valid"] = False

        except Exception as e:
            validation_result["errors"].append(str(e))
            validation_result["is_valid"] = False

        return validation_result

    def _looks_like_assumption(self, cell: Dict[str, Any]) -> bool:
        """Check if a cell looks like it contains an assumption."""
        if not isinstance(cell, dict):
            return False

        # Check if it's a numeric value
        value = cell.get("value")
        if value is None:
            return False

        try:
            float_val = float(value)
            # Assumptions are often percentages or small numbers
            return 0 <= abs(float_val) <= 10
        except (ValueError, TypeError):
            return False

    def _create_assumption_from_cell(
        self, cell: Dict[str, Any], sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """Create an assumption dictionary from a cell."""
        try:
            value = float(cell.get("value", 0))
            return {
                "name": f"{sheet_name}_{cell.get('address', 'unknown')}",
                "value": value,
                "sheet": sheet_name,
                "address": cell.get("address"),
                "confidence": 0.7,
            }
        except (ValueError, TypeError):
            return None

    def _categorize_assumption(self, assumption: Dict[str, Any]) -> str:
        """Categorize an assumption based on its name and value."""
        name = assumption.get("name", "").lower()

        if any(keyword in name for keyword in ["growth", "rate", "increase"]):
            return "growth"
        elif any(keyword in name for keyword in ["margin", "cost", "expense"]):
            return "operational"
        else:
            return "financial"
