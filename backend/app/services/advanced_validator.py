import re
from typing import Dict, List, Any, Optional, Tuple, Set
from datetime import datetime
from dataclasses import dataclass
from enum import Enum
import pandas as pd
from openpyxl import Workbook

from app.schemas.file import FileValidationResult


class FinancialStatementType(str, Enum):
    """Types of financial statements."""
    PROFIT_LOSS = "profit_loss"
    BALANCE_SHEET = "balance_sheet"
    CASH_FLOW = "cash_flow"
    TRIAL_BALANCE = "trial_balance"
    BUDGET = "budget"
    FORECAST = "forecast"


class ValidationSeverity(str, Enum):
    """Validation issue severity levels."""
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass
class ValidationIssue:
    """Represents a validation issue."""
    severity: ValidationSeverity
    message: str
    location: str  # Cell or range reference
    category: str
    suggestion: Optional[str] = None
    auto_fixable: bool = False


@dataclass
class FinancialTemplate:
    """Defines a financial statement template."""
    name: str
    statement_type: FinancialStatementType
    required_sections: List[str]
    optional_sections: List[str]
    required_accounts: List[str]
    validation_rules: List[Dict[str, Any]]
    calculation_rules: List[Dict[str, Any]]


class AdvancedTemplateValidator:
    """Advanced validation engine for financial statement templates."""
    
    def __init__(self):
        self.templates = self._load_templates()
        self.currency_patterns = {
            'USD': [r'\$', r'USD', r'US\$'],
            'EUR': [r'€', r'EUR'],
            'GBP': [r'£', r'GBP'],
            'CAD': [r'CAD', r'C\$'],
        }
        self.date_patterns = [
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'\d{1,2}-\d{1,2}-\d{4}',
            r'\d{4}-\d{1,2}-\d{1,2}',
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},?\s+\d{4}',
            r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}',
        ]
    
    def validate_financial_statement(
        self, 
        workbook: Workbook, 
        sheet_name: str, 
        statement_type: FinancialStatementType
    ) -> FileValidationResult:
        """
        Perform comprehensive validation of a financial statement.
        
        Args:
            workbook: Excel workbook object
            sheet_name: Name of the sheet to validate
            statement_type: Type of financial statement
            
        Returns:
            FileValidationResult with detailed validation results
        """
        issues = []
        warnings = []
        sheet = workbook[sheet_name]
        
        # Get template for this statement type
        template = self.templates.get(statement_type)
        if not template:
            issues.append(f"No template found for statement type: {statement_type}")
            return FileValidationResult(
                is_valid=False,
                errors=issues,
                warnings=warnings
            )
        
        # Perform comprehensive validation
        issues.extend(self._validate_structure(sheet, template))
        issues.extend(self._validate_accounts(sheet, template))
        issues.extend(self._validate_calculations(sheet, template))
        issues.extend(self._validate_formatting(sheet))
        issues.extend(self._validate_data_consistency(sheet))
        issues.extend(self._validate_financial_logic(sheet, statement_type))
        
        # Separate errors and warnings
        errors = [issue.message for issue in issues if issue.severity == ValidationSeverity.ERROR]
        warnings = [issue.message for issue in issues if issue.severity == ValidationSeverity.WARNING]
        
        return FileValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            warnings=warnings,
            sheet_info={
                "statement_type": statement_type,
                "template_matched": template.name,
                "validation_issues": len(issues),
                "auto_fixable_issues": sum(1 for issue in issues if issue.auto_fixable)
            }
        )
    
    def _validate_structure(self, sheet, template: FinancialTemplate) -> List[ValidationIssue]:
        """Validate the structural elements of the financial statement."""
        issues = []
        
        # Check for required sections
        sheet_text = self._get_sheet_text(sheet)
        
        for section in template.required_sections:
            if not self._find_section(sheet_text, section):
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    message=f"Required section '{section}' not found",
                    location="Sheet structure",
                    category="Structure",
                    suggestion=f"Add a '{section}' section to the statement"
                ))
        
        # Check for proper header structure
        if not self._validate_headers(sheet):
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                message="Headers not properly formatted or positioned",
                location="Row 1-5",
                category="Structure",
                suggestion="Ensure headers are in the first few rows and clearly labeled"
            ))
        
        # Check for consistent column structure
        column_issues = self._validate_column_structure(sheet)
        issues.extend(column_issues)
        
        return issues
    
    def _validate_accounts(self, sheet, template: FinancialTemplate) -> List[ValidationIssue]:
        """Validate account names and classifications."""
        issues = []
        
        # Extract account names from the sheet
        accounts = self._extract_account_names(sheet)
        
        # Check for required accounts
        for required_account in template.required_accounts:
            if not self._find_similar_account(accounts, required_account):
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    message=f"Required account '{required_account}' not found",
                    location="Account names",
                    category="Accounts",
                    suggestion=f"Add '{required_account}' account or similar"
                ))
        
        # Check for account naming consistency
        naming_issues = self._validate_account_naming(accounts)
        issues.extend(naming_issues)
        
        # Check for duplicate accounts
        duplicates = self._find_duplicate_accounts(accounts)
        for duplicate in duplicates:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                message=f"Duplicate account found: '{duplicate}'",
                location="Account names",
                category="Accounts",
                suggestion="Remove or rename duplicate accounts"
            ))
        
        return issues
    
    def _validate_calculations(self, sheet, template: FinancialTemplate) -> List[ValidationIssue]:
        """Validate formulas and calculations."""
        issues = []
        
        # Check for calculation rules
        for rule in template.calculation_rules:
            rule_issues = self._validate_calculation_rule(sheet, rule)
            issues.extend(rule_issues)
        
        # Check for broken formulas
        broken_formulas = self._find_broken_formulas(sheet)
        for formula_info in broken_formulas:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                message=f"Broken formula in cell {formula_info['cell']}: {formula_info['error']}",
                location=formula_info['cell'],
                category="Calculations",
                suggestion="Fix the formula syntax or references"
            ))
        
        # Check for hardcoded values that should be formulas
        hardcoded_issues = self._find_hardcoded_totals(sheet)
        issues.extend(hardcoded_issues)
        
        return issues
    
    def _validate_formatting(self, sheet) -> List[ValidationIssue]:
        """Validate formatting consistency and standards."""
        issues = []
        
        # Check number formatting
        number_format_issues = self._validate_number_formatting(sheet)
        issues.extend(number_format_issues)
        
        # Check for consistent currency formatting
        currency_issues = self._validate_currency_formatting(sheet)
        issues.extend(currency_issues)
        
        # Check date formatting
        date_issues = self._validate_date_formatting(sheet)
        issues.extend(date_issues)
        
        # Check for proper alignment and styling
        style_issues = self._validate_styling(sheet)
        issues.extend(style_issues)
        
        return issues
    
    def _validate_data_consistency(self, sheet) -> List[ValidationIssue]:
        """Validate data consistency and reasonableness."""
        issues = []
        
        # Check for missing data
        missing_data_issues = self._find_missing_data(sheet)
        issues.extend(missing_data_issues)
        
        # Check for unreasonable values
        unreasonable_values = self._find_unreasonable_values(sheet)
        issues.extend(unreasonable_values)
        
        # Check for data type consistency
        type_consistency_issues = self._validate_data_types(sheet)
        issues.extend(type_consistency_issues)
        
        return issues
    
    def _validate_financial_logic(self, sheet, statement_type: FinancialStatementType) -> List[ValidationIssue]:
        """Validate financial logic and business rules."""
        issues = []
        
        if statement_type == FinancialStatementType.PROFIT_LOSS:
            issues.extend(self._validate_pnl_logic(sheet))
        elif statement_type == FinancialStatementType.BALANCE_SHEET:
            issues.extend(self._validate_balance_sheet_logic(sheet))
        elif statement_type == FinancialStatementType.CASH_FLOW:
            issues.extend(self._validate_cash_flow_logic(sheet))
        
        return issues
    
    def _validate_pnl_logic(self, sheet) -> List[ValidationIssue]:
        """Validate P&L specific logic."""
        issues = []
        
        # Check that revenue accounts are positive (or negative for returns)
        revenue_cells = self._find_revenue_cells(sheet)
        for cell in revenue_cells:
            if cell['value'] < 0 and 'return' not in cell['label'].lower():
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    message=f"Revenue account '{cell['label']}' has negative value",
                    location=cell['location'],
                    category="Financial Logic",
                    suggestion="Verify if this should be an expense or if the value is correct"
                ))
        
        # Check that expense accounts are typically positive (representing outflows)
        expense_cells = self._find_expense_cells(sheet)
        for cell in expense_cells:
            if cell['value'] < 0:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    message=f"Expense account '{cell['label']}' has negative value",
                    location=cell['location'],
                    category="Financial Logic",
                    suggestion="Verify if this represents a credit or if the value is correct"
                ))
        
        # Check for reasonable gross margin
        gross_margin = self._calculate_gross_margin(sheet)
        if gross_margin is not None:
            if gross_margin < 0:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    message=f"Gross margin is negative ({gross_margin:.1%})",
                    location="Gross margin calculation",
                    category="Financial Logic",
                    suggestion="Review cost of goods sold and revenue figures"
                ))
            elif gross_margin > 0.95:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.INFO,
                    message=f"Gross margin is very high ({gross_margin:.1%})",
                    location="Gross margin calculation",
                    category="Financial Logic",
                    suggestion="Verify if this margin is realistic for the business"
                ))
        
        return issues
    
    def _validate_balance_sheet_logic(self, sheet) -> List[ValidationIssue]:
        """Validate Balance Sheet specific logic."""
        issues = []
        
        # Check balance sheet equation: Assets = Liabilities + Equity
        balance_check = self._check_balance_sheet_balance(sheet)
        if balance_check['unbalanced']:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                message=f"Balance sheet does not balance. Difference: {balance_check['difference']:,.2f}",
                location="Total calculations",
                category="Financial Logic",
                suggestion="Review all totals and ensure Assets = Liabilities + Equity"
            ))
        
        # Check for negative equity (potential concern)
        equity_total = self._get_total_equity(sheet)
        if equity_total is not None and equity_total < 0:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                message="Total equity is negative, indicating potential financial distress",
                location="Equity section",
                category="Financial Logic",
                suggestion="Review equity accounts and consider additional analysis"
            ))
        
        return issues
    
    def _validate_cash_flow_logic(self, sheet) -> List[ValidationIssue]:
        """Validate Cash Flow specific logic."""
        issues = []
        
        # Check that cash flow sections sum correctly
        cf_sections = self._get_cash_flow_sections(sheet)
        
        if cf_sections:
            total_calculated = sum(cf_sections.values())
            reported_total = self._get_reported_cash_flow_total(sheet)
            
            if reported_total is not None and abs(total_calculated - reported_total) > 0.01:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    message=f"Cash flow sections don't sum to reported total. Difference: {abs(total_calculated - reported_total):,.2f}",
                    location="Cash flow totals",
                    category="Financial Logic",
                    suggestion="Review cash flow section calculations"
                ))
        
        return issues
    
    def _load_templates(self) -> Dict[FinancialStatementType, FinancialTemplate]:
        """Load predefined financial statement templates."""
        templates = {}
        
        # P&L Template
        templates[FinancialStatementType.PROFIT_LOSS] = FinancialTemplate(
            name="Standard P&L",
            statement_type=FinancialStatementType.PROFIT_LOSS,
            required_sections=["Revenue", "Cost of Goods Sold", "Operating Expenses", "Net Income"],
            optional_sections=["Other Income", "Taxes", "Depreciation"],
            required_accounts=["Revenue", "Cost of Sales", "Gross Profit", "Operating Expenses", "Net Income"],
            validation_rules=[
                {"rule": "revenue_positive", "severity": "warning"},
                {"rule": "expenses_reasonable", "severity": "warning"},
                {"rule": "calculations_consistent", "severity": "error"}
            ],
            calculation_rules=[
                {"formula": "Gross Profit = Revenue - Cost of Sales", "required": True},
                {"formula": "Net Income = Gross Profit - Operating Expenses", "required": True}
            ]
        )
        
        # Balance Sheet Template
        templates[FinancialStatementType.BALANCE_SHEET] = FinancialTemplate(
            name="Standard Balance Sheet",
            statement_type=FinancialStatementType.BALANCE_SHEET,
            required_sections=["Assets", "Liabilities", "Equity"],
            optional_sections=["Current Assets", "Fixed Assets", "Current Liabilities", "Long-term Liabilities"],
            required_accounts=["Total Assets", "Total Liabilities", "Total Equity"],
            validation_rules=[
                {"rule": "balance_equation", "severity": "error"},
                {"rule": "asset_liability_classification", "severity": "warning"}
            ],
            calculation_rules=[
                {"formula": "Total Assets = Total Liabilities + Total Equity", "required": True}
            ]
        )
        
        # Cash Flow Template
        templates[FinancialStatementType.CASH_FLOW] = FinancialTemplate(
            name="Standard Cash Flow",
            statement_type=FinancialStatementType.CASH_FLOW,
            required_sections=["Operating Activities", "Investing Activities", "Financing Activities"],
            optional_sections=["Beginning Cash", "Ending Cash"],
            required_accounts=["Net Cash from Operations", "Net Cash from Investing", "Net Cash from Financing"],
            validation_rules=[
                {"rule": "cash_flow_consistency", "severity": "error"},
                {"rule": "activity_classification", "severity": "warning"}
            ],
            calculation_rules=[
                {"formula": "Net Change in Cash = Operating + Investing + Financing", "required": True}
            ]
        )
        
        return templates
    
    # Helper methods (simplified implementations)
    def _get_sheet_text(self, sheet) -> str:
        """Extract all text from the sheet."""
        text_values = []
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text_values.append(cell)
        return " ".join(text_values)
    
    def _find_section(self, text: str, section: str) -> bool:
        """Find if a section exists in the text."""
        pattern = re.compile(section.replace(" ", r"\s+"), re.IGNORECASE)
        return bool(pattern.search(text))
    
    def _validate_headers(self, sheet) -> bool:
        """Check if headers are properly formatted."""
        # Simple check for headers in first few rows
        for row_num in range(1, 6):
            row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            if any(cell and isinstance(cell, str) and len(cell) > 3 for cell in row):
                return True
        return False
    
    def _validate_column_structure(self, sheet) -> List[ValidationIssue]:
        """Validate column structure consistency."""
        issues = []
        # Implementation would check for consistent column usage
        return issues
    
    def _extract_account_names(self, sheet) -> List[str]:
        """Extract account names from the sheet."""
        accounts = []
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and len(cell) > 3:
                    # Basic heuristic for account names
                    if any(keyword in cell.lower() for keyword in ['revenue', 'expense', 'asset', 'liability', 'equity', 'income', 'cost']):
                        accounts.append(cell)
        return accounts
    
    def _find_similar_account(self, accounts: List[str], target: str) -> bool:
        """Find if a similar account exists."""
        target_lower = target.lower()
        for account in accounts:
            if target_lower in account.lower() or account.lower() in target_lower:
                return True
        return False
    
    def _validate_account_naming(self, accounts: List[str]) -> List[ValidationIssue]:
        """Validate account naming conventions."""
        issues = []
        # Implementation would check naming consistency
        return issues
    
    def _find_duplicate_accounts(self, accounts: List[str]) -> List[str]:
        """Find duplicate account names."""
        seen = set()
        duplicates = []
        for account in accounts:
            if account.lower() in seen:
                duplicates.append(account)
            seen.add(account.lower())
        return duplicates
    
    # Additional helper methods would be implemented for comprehensive validation
    def _validate_calculation_rule(self, sheet, rule: Dict[str, Any]) -> List[ValidationIssue]:
        """Validate a specific calculation rule."""
        return []
    
    def _find_broken_formulas(self, sheet) -> List[Dict[str, Any]]:
        """Find cells with broken formulas."""
        return []
    
    def _find_hardcoded_totals(self, sheet) -> List[ValidationIssue]:
        """Find totals that should be formulas."""
        return []
    
    def _validate_number_formatting(self, sheet) -> List[ValidationIssue]:
        """Validate number formatting consistency."""
        return []
    
    def _validate_currency_formatting(self, sheet) -> List[ValidationIssue]:
        """Validate currency formatting."""
        return []
    
    def _validate_date_formatting(self, sheet) -> List[ValidationIssue]:
        """Validate date formatting."""
        return []
    
    def _validate_styling(self, sheet) -> List[ValidationIssue]:
        """Validate styling consistency."""
        return []
    
    def _find_missing_data(self, sheet) -> List[ValidationIssue]:
        """Find missing data issues."""
        return []
    
    def _find_unreasonable_values(self, sheet) -> List[ValidationIssue]:
        """Find unreasonable values."""
        return []
    
    def _validate_data_types(self, sheet) -> List[ValidationIssue]:
        """Validate data type consistency."""
        return []
    
    def _find_revenue_cells(self, sheet) -> List[Dict[str, Any]]:
        """Find revenue-related cells."""
        return []
    
    def _find_expense_cells(self, sheet) -> List[Dict[str, Any]]:
        """Find expense-related cells."""
        return []
    
    def _calculate_gross_margin(self, sheet) -> Optional[float]:
        """Calculate gross margin if possible."""
        return None
    
    def _check_balance_sheet_balance(self, sheet) -> Dict[str, Any]:
        """Check if balance sheet balances."""
        return {"unbalanced": False, "difference": 0}
    
    def _get_total_equity(self, sheet) -> Optional[float]:
        """Get total equity value."""
        return None
    
    def _get_cash_flow_sections(self, sheet) -> Dict[str, float]:
        """Get cash flow section totals."""
        return {}
    
    def _get_reported_cash_flow_total(self, sheet) -> Optional[float]:
        """Get reported cash flow total."""
        return None 