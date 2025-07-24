import os
import re
import pandas as pd
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime, date
from pathlib import Path
from dataclasses import dataclass, field
from enum import Enum

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataType(str, Enum):
    """Enumeration of Excel data types."""
    NUMBER = "number"
    TEXT = "text"
    DATE = "date"
    BOOLEAN = "boolean"
    FORMULA = "formula"
    ERROR = "error"
    BLANK = "blank"


class SheetType(str, Enum):
    """Types of financial statement sheets."""
    PROFIT_LOSS = "profit_loss"
    BALANCE_SHEET = "balance_sheet"
    CASH_FLOW = "cash_flow"
    ASSUMPTIONS = "assumptions"
    CALCULATIONS = "calculations"
    SUMMARY = "summary"
    OTHER = "other"


@dataclass
class CellInfo:
    """Information about a single Excel cell."""
    address: str
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    data_type: DataType = DataType.BLANK
    number_format: Optional[str] = None
    is_merged: bool = False
    merged_range: Optional[str] = None
    has_comment: bool = False
    comment_text: Optional[str] = None


@dataclass
class SheetInfo:
    """Information about an Excel worksheet."""
    name: str
    sheet_type: SheetType
    max_row: int
    max_column: int
    data_range: Optional[str] = None
    header_row: Optional[int] = None
    data_start_row: Optional[int] = None
    has_formulas: bool = False
    formula_count: int = 0
    cells: List[CellInfo] = field(default_factory=list)
    financial_sections: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ValidationError:
    """Validation error information."""
    severity: str  # error, warning, info
    message: str
    sheet: Optional[str] = None
    cell: Optional[str] = None
    suggestion: Optional[str] = None


@dataclass
class ValidationSummary:
    """Summary of validation results."""
    is_valid: bool
    errors: List[ValidationError] = field(default_factory=list)
    warnings: List[ValidationError] = field(default_factory=list)
    total_errors: int = 0
    total_warnings: int = 0


@dataclass
class ParsedData:
    """Complete parsed Excel data structure."""
    file_name: str
    file_path: str
    file_size: int
    sheets: List[SheetInfo] = field(default_factory=list)
    validation_summary: ValidationSummary = field(default_factory=ValidationSummary)
    metadata: Dict[str, Any] = field(default_factory=dict)
    formulas: Dict[str, str] = field(default_factory=dict)
    dependencies: Dict[str, List[str]] = field(default_factory=dict)
    time_series_data: Dict[str, Any] = field(default_factory=dict)
    financial_metrics: Dict[str, Any] = field(default_factory=dict)


class ExcelParser:
    """Advanced Excel file parser for financial models."""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel parsing")
        
        # Financial statement patterns
        self.pl_patterns = [
            r'(profit|loss|income|statement|p&l|pnl)',
            r'(revenue|sales|income)',
            r'(expense|cost|expenditure)',
            r'(ebitda|ebit|operating)',
            r'(net\s+income|net\s+profit)'
        ]
        
        self.balance_sheet_patterns = [
            r'(balance|sheet|position)',
            r'(assets|liabilities|equity)',
            r'(current|non.current|fixed)',
            r'(cash|receivables|payables)',
            r'(debt|equity|retained)'
        ]
        
        self.cash_flow_patterns = [
            r'(cash|flow|cf)',
            r'(operating|investing|financing)',
            r'(capex|capital|expenditure)',
            r'(depreciation|amortization)',
            r'(working\s+capital)'
        ]
        
        # Date patterns
        self.date_patterns = [
            r'\d{4}',  # Year
            r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',
            r'(q1|q2|q3|q4|quarter)',
            r'(fy|year|annual)',
            r'\d{1,2}/\d{1,2}/\d{2,4}'  # Date format
        ]

    def parse_file(self, file_path: str) -> ParsedData:
        """Alias for parse_excel_file for backward compatibility."""
        return self.parse_excel_file(file_path)

    def parse_excel_file(self, file_path: str) -> ParsedData:
        """
        Parse Excel file and extract comprehensive data.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ParsedData object containing all extracted information
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        file_info = Path(file_path)
        
        try:
            # Load workbook with formulas
            workbook = load_workbook(file_path, data_only=False)
            
            # Initialize parsed data
            parsed_data = ParsedData(
                file_name=file_info.name,
                file_path=str(file_path),
                file_size=file_info.stat().st_size
            )
            
            # Extract metadata
            parsed_data.metadata = self._extract_metadata(workbook)
            
            # Parse each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = self._parse_worksheet(sheet)
                parsed_data.sheets.append(sheet_info)
            
            # Extract formulas and dependencies
            parsed_data.formulas, parsed_data.dependencies = self._extract_formulas(workbook)
            
            # Extract time series data
            parsed_data.time_series_data = self._extract_time_series(parsed_data.sheets)
            
            # Calculate financial metrics
            parsed_data.financial_metrics = self._calculate_basic_metrics(parsed_data.sheets)
            
            # Validate data
            parsed_data.validation_summary = self._validate_data(parsed_data)
            
            return parsed_data
            
        except Exception as e:
            # Create error response
            error_data = ParsedData(
                file_name=file_info.name,
                file_path=str(file_path),
                file_size=file_info.stat().st_size if file_info.exists() else 0
            )
            
            error_data.validation_summary = ValidationSummary(
                is_valid=False,
                errors=[ValidationError(
                    severity="error",
                    message=f"Failed to parse Excel file: {str(e)}",
                    suggestion="Please check if the file is a valid Excel file (.xlsx) and is not corrupted"
                )],
                total_errors=1
            )
            
            return error_data

    def _extract_metadata(self, workbook: Workbook) -> Dict[str, Any]:
        """Extract file metadata."""
        properties = workbook.properties
        return {
            "title": properties.title,
            "subject": properties.subject,
            "creator": properties.creator,
            "created": properties.created.isoformat() if properties.created else None,
            "modified": properties.modified.isoformat() if properties.modified else None,
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames
        }

    def _parse_worksheet(self, sheet: Worksheet) -> SheetInfo:
        """Parse individual worksheet."""
        sheet_info = SheetInfo(
            name=sheet.title,
            sheet_type=self._detect_sheet_type(sheet),
            max_row=sheet.max_row,
            max_column=sheet.max_column
        )
        
        # Find data range
        sheet_info.data_range = self._find_data_range(sheet)
        
        # Detect header row and data start
        sheet_info.header_row, sheet_info.data_start_row = self._detect_headers(sheet)
        
        # Parse cells
        formula_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.formula:
                    cell_info = self._parse_cell(cell)
                    sheet_info.cells.append(cell_info)
                    
                    if cell_info.data_type == DataType.FORMULA:
                        formula_count += 1
        
        sheet_info.has_formulas = formula_count > 0
        sheet_info.formula_count = formula_count
        
        # Identify financial sections
        sheet_info.financial_sections = self._identify_financial_sections(sheet)
        
        return sheet_info

    def _parse_cell(self, cell: Cell) -> CellInfo:
        """Parse individual cell."""
        cell_info = CellInfo(
            address=cell.coordinate,
            row=cell.row,
            column=cell.column,
            value=cell.value
        )
        
        # Handle formulas
        if cell.formula:
            cell_info.formula = cell.formula
            cell_info.data_type = DataType.FORMULA
        else:
            cell_info.data_type = self._detect_data_type(cell.value)
        
        # Number format
        cell_info.number_format = cell.number_format
        
        # Check if merged
        if hasattr(cell, 'merged_cells') and cell.coordinate in cell.parent.merged_cells:
            cell_info.is_merged = True
            # Find merged range
            for merged_range in cell.parent.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell_info.merged_range = str(merged_range)
                    break
        
        # Comments
        if cell.comment:
            cell_info.has_comment = True
            cell_info.comment_text = cell.comment.text
        
        return cell_info

    def _detect_data_type(self, value: Any) -> DataType:
        """Detect the data type of a cell value."""
        if value is None:
            return DataType.BLANK
        elif isinstance(value, bool):
            return DataType.BOOLEAN
        elif isinstance(value, (int, float)):
            return DataType.NUMBER
        elif isinstance(value, (date, datetime)):
            return DataType.DATE
        elif isinstance(value, str):
            # Check for error values
            if value.startswith('#'):
                return DataType.ERROR
            return DataType.TEXT
        else:
            return DataType.TEXT

    def _detect_sheet_type(self, sheet: Worksheet) -> SheetType:
        """Detect the type of financial statement sheet."""
        sheet_name = sheet.title.lower()
        
        # Get text content from first few rows to analyze
        text_content = ""
        for row in range(1, min(10, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    text_content += " " + cell.value.lower()
        
        combined_text = sheet_name + " " + text_content
        
        # Check patterns
        pl_score = sum(1 for pattern in self.pl_patterns 
                      if re.search(pattern, combined_text, re.IGNORECASE))
        bs_score = sum(1 for pattern in self.balance_sheet_patterns 
                      if re.search(pattern, combined_text, re.IGNORECASE))
        cf_score = sum(1 for pattern in self.cash_flow_patterns 
                      if re.search(pattern, combined_text, re.IGNORECASE))
        
        # Determine sheet type
        max_score = max(pl_score, bs_score, cf_score)
        if max_score == 0:
            return SheetType.OTHER
        elif pl_score == max_score:
            return SheetType.PROFIT_LOSS
        elif bs_score == max_score:
            return SheetType.BALANCE_SHEET
        elif cf_score == max_score:
            return SheetType.CASH_FLOW
        else:
            return SheetType.OTHER

    def _find_data_range(self, sheet: Worksheet) -> Optional[str]:
        """Find the actual data range in the sheet."""
        min_row, max_row = 1, sheet.max_row
        min_col, max_col = 1, sheet.max_column
        
        # Find first non-empty cell
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    min_row = min(min_row, row)
                    max_row = max(max_row, row)
                    min_col = min(min_col, col)
                    max_col = max(max_col, col)
        
        if min_row <= max_row and min_col <= max_col:
            start_cell = f"{get_column_letter(min_col)}{min_row}"
            end_cell = f"{get_column_letter(max_col)}{max_row}"
            return f"{start_cell}:{end_cell}"
        
        return None

    def _detect_headers(self, sheet: Worksheet) -> Tuple[Optional[int], Optional[int]]:
        """Detect header row and data start row."""
        header_row = None
        data_start_row = None
        
        # Look for header patterns in first 10 rows
        for row in range(1, min(11, sheet.max_row + 1)):
            row_text = ""
            for col in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    row_text += " " + cell.value.lower()
            
            # Check if this looks like a header row
            if any(keyword in row_text for keyword in 
                   ['account', 'description', 'amount', 'value', 'period', 'year', 'month']):
                header_row = row
                data_start_row = row + 1
                break
        
        return header_row, data_start_row

    def _extract_formulas(self, workbook: Workbook) -> Tuple[Dict[str, str], Dict[str, List[str]]]:
        """Extract formulas and their dependencies."""
        formulas = {}
        dependencies = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.formula:
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        formulas[cell_ref] = cell.formula
                        
                        # Extract dependencies from formula
                        deps = self._extract_cell_references(cell.formula)
                        if deps:
                            dependencies[cell_ref] = deps
        
        return formulas, dependencies

    def _extract_cell_references(self, formula: str) -> List[str]:
        """Extract cell references from a formula."""
        # Simplified regex for cell references
        pattern = r'[A-Z]+\d+'
        references = re.findall(pattern, formula.upper())
        return list(set(references))  # Remove duplicates

    def _identify_financial_sections(self, sheet: Worksheet) -> Dict[str, Any]:
        """Identify financial sections in the sheet."""
        sections = {}
        
        # This is a simplified implementation
        # In practice, you'd want more sophisticated pattern matching
        for row in range(1, min(50, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    text = cell.value.lower()
                    
                    # Look for section headers
                    if 'revenue' in text or 'sales' in text:
                        sections['revenue_section'] = {'row': row, 'col': col}
                    elif 'expense' in text or 'cost' in text:
                        sections['expense_section'] = {'row': row, 'col': col}
                    elif 'total' in text:
                        sections['total_section'] = {'row': row, 'col': col}
        
        return sections

    def _extract_time_series(self, sheets: List[SheetInfo]) -> Dict[str, Any]:
        """Extract time series data from sheets."""
        time_series = {}
        
        for sheet in sheets:
            # Look for date patterns in the first few rows
            date_columns = []
            for cell in sheet.cells[:50]:  # Check first 50 cells
                if cell.data_type == DataType.DATE or self._looks_like_date(cell.value):
                    date_columns.append(cell.column)
            
            if date_columns:
                time_series[sheet.name] = {
                    'date_columns': date_columns,
                    'has_time_series': True
                }
        
        return time_series

    def _looks_like_date(self, value: Any) -> bool:
        """Check if a value looks like a date."""
        if not isinstance(value, str):
            return False
        
        return any(re.search(pattern, value.lower()) for pattern in self.date_patterns)

    def _calculate_basic_metrics(self, sheets: List[SheetInfo]) -> Dict[str, Any]:
        """Calculate basic financial metrics from the data."""
        metrics = {}
        
        for sheet in sheets:
            sheet_metrics = {
                'total_cells': len(sheet.cells),
                'formula_cells': sheet.formula_count,
                'data_density': len(sheet.cells) / (sheet.max_row * sheet.max_column) if sheet.max_row > 0 and sheet.max_column > 0 else 0,
                'has_financial_data': len(sheet.financial_sections) > 0
            }
            
            # Count data types
            type_counts = {}
            for cell in sheet.cells:
                data_type = cell.data_type.value
                type_counts[data_type] = type_counts.get(data_type, 0) + 1
            
            sheet_metrics['data_type_distribution'] = type_counts
            metrics[sheet.name] = sheet_metrics
        
        return metrics

    def _validate_data(self, parsed_data: ParsedData) -> ValidationSummary:
        """Validate the parsed Excel data."""
        errors = []
        warnings = []
        
        # Check if file has any data
        if not parsed_data.sheets:
            errors.append(ValidationError(
                severity="error",
                message="No worksheets found in the Excel file",
                suggestion="Please ensure the file contains at least one worksheet with data"
            ))
        
        # Check for financial statement sheets
        financial_sheets = [s for s in parsed_data.sheets 
                          if s.sheet_type in [SheetType.PROFIT_LOSS, SheetType.BALANCE_SHEET, SheetType.CASH_FLOW]]
        
        if not financial_sheets:
            warnings.append(ValidationError(
                severity="warning",
                message="No recognized financial statement sheets found",
                suggestion="Consider renaming sheets to include keywords like 'P&L', 'Balance Sheet', or 'Cash Flow'"
            ))
        
        # Check for formulas
        total_formulas = sum(sheet.formula_count for sheet in parsed_data.sheets)
        if total_formulas == 0:
            warnings.append(ValidationError(
                severity="warning",
                message="No formulas found in the Excel file",
                suggestion="Financial models typically contain formulas for calculations"
            ))
        
        # Validate individual sheets
        for sheet in parsed_data.sheets:
            if sheet.max_row == 0 or sheet.max_column == 0:
                warnings.append(ValidationError(
                    severity="warning",
                    message=f"Sheet '{sheet.name}' appears to be empty",
                    sheet=sheet.name,
                    suggestion="Consider removing empty sheets or adding data"
                ))
        
        return ValidationSummary(
            is_valid=len(errors) == 0,
            errors=errors,
            warnings=warnings,
            total_errors=len(errors),
            total_warnings=len(warnings)
        )

    def export_to_dict(self, parsed_data: ParsedData) -> Dict[str, Any]:
        """Export parsed data to dictionary format."""
        return {
            "file_info": {
                "name": parsed_data.file_name,
                "path": parsed_data.file_path,
                "size": parsed_data.file_size
            },
            "metadata": parsed_data.metadata,
            "sheets": [
                {
                    "name": sheet.name,
                    "type": sheet.sheet_type.value,
                    "dimensions": {
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "data_range": sheet.data_range
                    },
                    "structure": {
                        "header_row": sheet.header_row,
                        "data_start_row": sheet.data_start_row,
                        "has_formulas": sheet.has_formulas,
                        "formula_count": sheet.formula_count
                    },
                    "financial_sections": sheet.financial_sections
                }
                for sheet in parsed_data.sheets
            ],
            "formulas": parsed_data.formulas,
            "dependencies": parsed_data.dependencies,
            "time_series": parsed_data.time_series_data,
            "metrics": parsed_data.financial_metrics,
            "validation": {
                "is_valid": parsed_data.validation_summary.is_valid,
                "errors": [
                    {
                        "severity": error.severity,
                        "message": error.message,
                        "sheet": error.sheet,
                        "cell": error.cell,
                        "suggestion": error.suggestion
                    }
                    for error in parsed_data.validation_summary.errors
                ],
                "warnings": [
                    {
                        "severity": warning.severity,
                        "message": warning.message,
                        "sheet": warning.sheet,
                        "cell": warning.cell,
                        "suggestion": warning.suggestion
                    }
                    for warning in parsed_data.validation_summary.warnings
                ]
            }
        }
