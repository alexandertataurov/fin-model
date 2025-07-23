import os
import io
import json
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from xlsxwriter import Workbook as XlsxWriterWorkbook

from app.core.config import settings


class ExcelExporter:
    """Service for exporting financial data to Excel with formula preservation."""
    
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir) if output_dir else Path(settings.UPLOAD_DIR) / "exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Excel styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.currency_format = '"$"#,##0.00_);("$"#,##0.00)'
        self.number_format = '#,##0.00'
        self.percentage_format = '0.00%'
        
        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_financial_data(
        self,
        data: Dict[str, Any],
        template_config: Optional[Dict[str, Any]] = None,
        preserve_formulas: bool = True,
        filename: Optional[str] = None
    ) -> str:
        """
        Export financial data to Excel with multiple sheets and preserved formatting.
        
        Args:
            data: Financial data organized by sections/sheets
            template_config: Template configuration for styling
            preserve_formulas: Whether to preserve Excel formulas
            filename: Output filename
            
        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"financial_export_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets based on data structure
        sheet_order = ['Summary', 'P&L', 'Balance Sheet', 'Cash Flow', 'Raw Data']
        
        for sheet_name in sheet_order:
            if sheet_name.lower().replace(' ', '_').replace('&', '') in data:
                sheet_data = data[sheet_name.lower().replace(' ', '_').replace('&', '')]
                self._create_sheet(wb, sheet_name, sheet_data, preserve_formulas)
        
        # Add charts sheet if chart data exists
        if 'charts' in data:
            self._create_charts_sheet(wb, data['charts'])
        
        # Add metadata sheet
        self._create_metadata_sheet(wb, data.get('metadata', {}))
        
        # Save workbook
        wb.save(output_path)
        
        return str(output_path)
    
    def _create_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        data: Dict[str, Any],
        preserve_formulas: bool = True
    ):
        """Create a worksheet with financial data."""
        ws = workbook.create_sheet(title=sheet_name)
        
        current_row = 1
        
        # Add sheet title
        ws.cell(row=current_row, column=1, value=f"{sheet_name} Analysis")
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(size=16, bold=True)
        current_row += 2
        
        # Add summary metrics if available
        if 'metrics' in data:
            current_row = self._add_metrics_section(ws, data['metrics'], current_row)
        
        # Add time series data if available
        if 'time_series' in data:
            current_row = self._add_time_series_section(ws, data['time_series'], current_row)
        
        # Add detailed data tables
        if 'tables' in data:
            for table_name, table_data in data['tables'].items():
                current_row = self._add_table_section(
                    ws, table_name, table_data, current_row, preserve_formulas
                )
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _add_metrics_section(
        self,
        worksheet,
        metrics_data: Dict[str, Any],
        start_row: int
    ) -> int:
        """Add key metrics section to worksheet."""
        current_row = start_row
        
        # Section header
        worksheet.cell(row=current_row, column=1, value="Key Metrics")
        header_cell = worksheet.cell(row=current_row, column=1)
        header_cell.font = Font(size=14, bold=True)
        current_row += 1
        
        # Create metrics table
        if isinstance(metrics_data, dict):
            # Headers
            worksheet.cell(row=current_row, column=1, value="Metric")
            worksheet.cell(row=current_row, column=2, value="Value")
            worksheet.cell(row=current_row, column=3, value="Previous Period")
            worksheet.cell(row=current_row, column=4, value="Change %")
            
            # Style headers
            for col in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
            
            current_row += 1
            
            # Add metrics data
            for metric_name, metric_value in metrics_data.items():
                worksheet.cell(row=current_row, column=1, value=metric_name.replace('_', ' ').title())
                
                if isinstance(metric_value, dict):
                    # Handle structured metric data
                    worksheet.cell(row=current_row, column=2, value=metric_value.get('current', ''))
                    worksheet.cell(row=current_row, column=3, value=metric_value.get('previous', ''))
                    
                    # Calculate percentage change
                    if metric_value.get('current') and metric_value.get('previous'):
                        try:
                            current_val = float(metric_value['current'])
                            previous_val = float(metric_value['previous'])
                            change_pct = ((current_val - previous_val) / previous_val) * 100
                            worksheet.cell(row=current_row, column=4, value=change_pct / 100)
                            worksheet.cell(row=current_row, column=4).number_format = self.percentage_format
                        except (ValueError, ZeroDivisionError):
                            pass
                else:
                    # Simple metric value
                    worksheet.cell(row=current_row, column=2, value=metric_value)
                
                # Apply formatting based on metric type
                if any(keyword in metric_name.lower() for keyword in ['revenue', 'profit', 'cost', 'expense', 'cash']):
                    worksheet.cell(row=current_row, column=2).number_format = self.currency_format
                    worksheet.cell(row=current_row, column=3).number_format = self.currency_format
                
                # Apply borders
                for col in range(1, 5):
                    worksheet.cell(row=current_row, column=col).border = self.thin_border
                
                current_row += 1
        
        return current_row + 2
    
    def _add_time_series_section(
        self,
        worksheet,
        time_series_data: List[Dict[str, Any]],
        start_row: int
    ) -> int:
        """Add time series data section to worksheet."""
        if not time_series_data:
            return start_row
        
        current_row = start_row
        
        # Section header
        worksheet.cell(row=current_row, column=1, value="Time Series Data")
        header_cell = worksheet.cell(row=current_row, column=1)
        header_cell.font = Font(size=14, bold=True)
        current_row += 1
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(time_series_data)
        
        # Add headers
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=column_name.replace('_', ' ').title())
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        current_row += 1
        start_data_row = current_row
        
        # Add data
        for _, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.thin_border
                
                # Apply number formatting
                if col_idx > 1 and isinstance(value, (int, float)):  # Skip date column
                    cell.number_format = self.currency_format
            
            current_row += 1
        
        # Create Excel table
        if len(df) > 0:
            table_range = f"A{start_data_row-1}:{chr(64+len(df.columns))}{current_row-1}"
            table = Table(displayName=f"TimeSeriesData", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
        
        return current_row + 2
    
    def _add_table_section(
        self,
        worksheet,
        table_name: str,
        table_data: List[Dict[str, Any]],
        start_row: int,
        preserve_formulas: bool = True
    ) -> int:
        """Add a data table section to worksheet."""
        if not table_data:
            return start_row
        
        current_row = start_row
        
        # Section header
        worksheet.cell(row=current_row, column=1, value=table_name.replace('_', ' ').title())
        header_cell = worksheet.cell(row=current_row, column=1)
        header_cell.font = Font(size=14, bold=True)
        current_row += 1
        
        # Convert to DataFrame
        df = pd.DataFrame(table_data)
        
        # Add to worksheet using openpyxl
        for r in dataframe_to_rows(df, index=False, header=True):
            for col_idx, value in enumerate(r, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                
                # Style header row
                if current_row == start_row + 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                
                cell.border = self.thin_border
                
                # Apply number formatting for data rows
                if current_row > start_row + 1 and isinstance(value, (int, float)):
                    if any(keyword in str(df.columns[col_idx-1]).lower() 
                          for keyword in ['revenue', 'profit', 'cost', 'expense', 'cash', 'amount']):
                        cell.number_format = self.currency_format
                    elif 'percentage' in str(df.columns[col_idx-1]).lower() or 'rate' in str(df.columns[col_idx-1]).lower():
                        cell.number_format = self.percentage_format
                    else:
                        cell.number_format = self.number_format
            
            current_row += 1
        
        return current_row + 2
    
    def _create_charts_sheet(self, workbook: Workbook, charts_data: Dict[str, Any]):
        """Create a sheet with Excel charts."""
        ws = workbook.create_sheet(title="Charts")
        
        current_row = 1
        ws.cell(row=current_row, column=1, value="Financial Charts")
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(size=16, bold=True)
        current_row += 3
        
        chart_row = current_row
        chart_col = 1
        
        for chart_name, chart_data in charts_data.items():
            if not chart_data:
                continue
            
            # Add chart title
            ws.cell(row=chart_row, column=chart_col, value=chart_name.replace('_', ' ').title())
            chart_row += 1
            
            # Create chart based on type
            if 'trend' in chart_name or 'time' in chart_name:
                chart = self._create_line_chart(ws, chart_data, chart_row, chart_col)
            elif 'breakdown' in chart_name or 'distribution' in chart_name:
                chart = self._create_pie_chart(ws, chart_data, chart_row, chart_col)
            else:
                chart = self._create_bar_chart(ws, chart_data, chart_row, chart_col)
            
            if chart:
                ws.add_chart(chart, f"{chr(65 + chart_col)}{chart_row}")
            
            chart_row += 20  # Space for next chart
            if chart_row > 60:  # Move to next column
                chart_row = current_row
                chart_col += 10
    
    def _create_line_chart(self, worksheet, data: List[Dict], start_row: int, start_col: int):
        """Create a line chart in Excel."""
        try:
            # Add data to worksheet
            df = pd.DataFrame(data)
            if df.empty or 'period' not in df.columns or 'value' not in df.columns:
                return None
            
            # Write data starting from start_row + 1
            data_start_row = start_row + 1
            for idx, row in df.iterrows():
                worksheet.cell(row=data_start_row + idx, column=start_col, value=row['period'])
                worksheet.cell(row=data_start_row + idx, column=start_col + 1, value=row['value'])
            
            # Create chart
            chart = LineChart()
            chart.title = "Trend Analysis"
            chart.style = 13
            chart.x_axis.title = 'Period'
            chart.y_axis.title = 'Value'
            
            # Data references
            data_ref = Reference(worksheet, min_col=start_col + 1, min_row=data_start_row,
                               max_row=data_start_row + len(df) - 1)
            categories_ref = Reference(worksheet, min_col=start_col, min_row=data_start_row,
                                     max_row=data_start_row + len(df) - 1)
            
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(categories_ref)
            
            return chart
        except Exception as e:
            print(f"Error creating line chart: {e}")
            return None
    
    def _create_bar_chart(self, worksheet, data: List[Dict], start_row: int, start_col: int):
        """Create a bar chart in Excel."""
        try:
            df = pd.DataFrame(data)
            if df.empty or 'period' not in df.columns or 'value' not in df.columns:
                return None
            
            # Write data
            data_start_row = start_row + 1
            for idx, row in df.iterrows():
                worksheet.cell(row=data_start_row + idx, column=start_col, value=row['period'])
                worksheet.cell(row=data_start_row + idx, column=start_col + 1, value=row['value'])
            
            # Create chart
            chart = BarChart()
            chart.title = "Category Analysis"
            chart.style = 10
            chart.x_axis.title = 'Category'
            chart.y_axis.title = 'Value'
            
            # Data references
            data_ref = Reference(worksheet, min_col=start_col + 1, min_row=data_start_row,
                               max_row=data_start_row + len(df) - 1)
            categories_ref = Reference(worksheet, min_col=start_col, min_row=data_start_row,
                                     max_row=data_start_row + len(df) - 1)
            
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(categories_ref)
            
            return chart
        except Exception as e:
            print(f"Error creating bar chart: {e}")
            return None
    
    def _create_pie_chart(self, worksheet, data: List[Dict], start_row: int, start_col: int):
        """Create a pie chart in Excel."""
        try:
            df = pd.DataFrame(data)
            if df.empty or 'period' not in df.columns or 'value' not in df.columns:
                return None
            
            # Write data
            data_start_row = start_row + 1
            for idx, row in df.iterrows():
                worksheet.cell(row=data_start_row + idx, column=start_col, value=row['period'])
                worksheet.cell(row=data_start_row + idx, column=start_col + 1, value=row['value'])
            
            # Create chart
            chart = PieChart()
            chart.title = "Distribution Analysis"
            
            # Data references
            data_ref = Reference(worksheet, min_col=start_col + 1, min_row=data_start_row,
                               max_row=data_start_row + len(df) - 1)
            categories_ref = Reference(worksheet, min_col=start_col, min_row=data_start_row,
                                     max_row=data_start_row + len(df) - 1)
            
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(categories_ref)
            
            return chart
        except Exception as e:
            print(f"Error creating pie chart: {e}")
            return None
    
    def _create_metadata_sheet(self, workbook: Workbook, metadata: Dict[str, Any]):
        """Create a metadata sheet with export information."""
        ws = workbook.create_sheet(title="Metadata")
        
        current_row = 1
        
        # Title
        ws.cell(row=current_row, column=1, value="Export Metadata")
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(size=16, bold=True)
        current_row += 2
        
        # Export information
        export_info = {
            "Export Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Generated By": "FinVision Financial Analysis System",
            "Export Type": "Financial Data Export",
            **metadata
        }
        
        for key, value in export_info.items():
            ws.cell(row=current_row, column=1, value=key)
            ws.cell(row=current_row, column=2, value=str(value))
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_raw_data_csv(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        include_metadata: bool = True
    ) -> str:
        """
        Export raw data to CSV format.
        
        Args:
            data: List of data records
            filename: Output filename
            include_metadata: Whether to include metadata header
            
        Returns:
            Path to generated CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_export_{timestamp}.csv"
        
        output_path = self.output_dir / filename
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Write CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            if include_metadata:
                f.write(f"# FinVision Data Export\n")
                f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"# Records: {len(df)}\n")
                f.write("\n")
            
            df.to_csv(f, index=False)
        
        return str(output_path) 